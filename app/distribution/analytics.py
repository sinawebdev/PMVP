"""Delivery analytics + export reports (Phase 4, Slice 3).

Analytics aggregates the same filtered delivery set the history view uses, grouped
by channel, status, and company, with success/failure rates. Exports stream that
filtered set as CSV or XLSX. Read-only, operator-plane, reusing
history.filtered_delivery_query so filters mean the same thing everywhere.
"""
import csv
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models import (
    DELIVERY_FAILED,
    DELIVERY_SENT,
    ClientCompany,
    PayrollRun,
    PayslipDelivery,
)

from .history import export_rows, filtered_delivery_query

EXPORT_COLUMNS = [
    "When", "Company", "Run", "Staff ID", "Employee", "Recipient",
    "Channel", "Status", "Attempts", "Initiated by", "Latest error",
]


def _pct(part, whole):
    return round(100 * part / whole, 1) if whole else 0.0


def _grouped(base, key_column):
    """{(key, status): count} for the filtered set grouped by a key + status."""
    from sqlalchemy import func

    rows = (
        base.with_entities(key_column, PayslipDelivery.status, func.count())
        .group_by(key_column, PayslipDelivery.status)
        .all()
    )
    return rows


def _summarise(rows_by_key):
    """[{key, total, sent, failed, success_rate, failure_rate}] sorted by volume."""
    out = []
    for key, buckets in rows_by_key.items():
        total = sum(buckets.values())
        sent = buckets.get(DELIVERY_SENT, 0)
        failed = buckets.get(DELIVERY_FAILED, 0)
        out.append({
            "key": key,
            "total": total,
            "sent": sent,
            "failed": failed,
            "success_rate": _pct(sent, sent + failed),
            "failure_rate": _pct(failed, sent + failed),
        })
    return sorted(out, key=lambda r: r["total"], reverse=True)


def delivery_analytics(filters):
    """Aggregate stats over the filtered delivery set: by channel, by company,
    plus overall totals — for the analytics page."""
    base = filtered_delivery_query(filters)

    by_channel_rows = _grouped(base, PayslipDelivery.channel)
    channels = {}
    for channel, status, count in by_channel_rows:
        channels.setdefault(channel, {})[status] = count

    by_company_rows = _grouped(base, PayrollRun.client_company_id)
    companies = {}
    for company_id, status, count in by_company_rows:
        companies.setdefault(company_id, {})[status] = count

    # Resolve company ids to names in one query.
    company_names = {}
    if companies:
        for c in ClientCompany.query.filter(ClientCompany.id.in_(companies.keys())).all():
            company_names[c.id] = c.name

    channel_summary = _summarise(channels)
    company_summary = _summarise(companies)
    for row in company_summary:
        row["name"] = company_names.get(row["key"], f"Company #{row['key']}")

    totals = {"total": 0, "sent": 0, "failed": 0}
    for row in channel_summary:
        totals["total"] += row["total"]
        totals["sent"] += row["sent"]
        totals["failed"] += row["failed"]
    totals["success_rate"] = _pct(totals["sent"], totals["sent"] + totals["failed"])
    totals["failure_rate"] = _pct(totals["failed"], totals["sent"] + totals["failed"])

    return {
        "totals": totals,
        "by_channel": channel_summary,
        "by_company": company_summary,
    }


def _row_values(d):
    run = d.payroll_run
    item = d.payroll_item
    batch = d.distribution_batch
    return [
        d.updated_at.strftime("%Y-%m-%d %H:%M") if d.updated_at else "",
        run.client_company.name if run and run.client_company else "",
        f"{run.month} {run.year}" if run else f"#{d.payroll_run_id}",
        item.staff_id if item else "",
        item.full_name if item else "",
        d.recipient or "",
        d.channel,
        d.status,
        d.attempts,
        (batch.initiated_by.name if batch and batch.initiated_by else "System") if batch else "",
        d.error or "",
    ]


def _timestamp():
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")


def export_deliveries_csv(filters):
    """(bytes, filename) — the filtered delivery set as CSV."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(EXPORT_COLUMNS)
    for d in export_rows(filters):
        writer.writerow(_row_values(d))
    data = buffer.getvalue().encode("utf-8-sig")  # BOM so Excel reads UTF-8
    return data, f"distribution-history-{_timestamp()}.csv"


def export_deliveries_xlsx(filters):
    """(bytes, filename) — the filtered set as XLSX with a summary sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Deliveries"
    ws.append(EXPORT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for d in export_rows(filters):
        ws.append(_row_values(d))

    summary = wb.create_sheet("Summary")
    stats = delivery_analytics(filters)
    summary.append(["Overall", "Total", "Sent", "Failed", "Success %", "Failure %"])
    summary["A1"].font = Font(bold=True)
    t = stats["totals"]
    summary.append(["All", t["total"], t["sent"], t["failed"], t["success_rate"], t["failure_rate"]])
    summary.append([])
    summary.append(["By channel", "Total", "Sent", "Failed", "Success %", "Failure %"])
    for row in stats["by_channel"]:
        summary.append([row["key"], row["total"], row["sent"], row["failed"],
                        row["success_rate"], row["failure_rate"]])
    summary.append([])
    summary.append(["By company", "Total", "Sent", "Failed", "Success %", "Failure %"])
    for row in stats["by_company"]:
        summary.append([row["name"], row["total"], row["sent"], row["failed"],
                        row["success_rate"], row["failure_rate"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), f"distribution-history-{_timestamp()}.xlsx"
