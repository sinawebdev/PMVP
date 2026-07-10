"""Generate the per-company monthly thin template for download.

The template is what the client fills in each month: one row per seeded worker,
the company's exact seeded hour-element columns, the adjustment columns
zero-filled, and a read-only **ICU Member** column scoped to members (so the
operator can see membership — ICU dues themselves are derived, never uploaded).
It parses straight back through the Phase 3 thin pipeline (`parse_thin_workbook`).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

from app.models import Employee, WageRateProfile
from app.excel_utils import save_workbook, slug_filename
from app.raw_engine.mapping import ELEMENT_SET
from app.raw_engine.thin import thin_header

_LABEL_BY_CODE = {code: label for code, label, _cat in ELEMENT_SET}
ICU_MEMBER_COLUMN = "ICU Member"


def seeded_element_codes(client_company_id):
    """The company's seeded pay-element codes, ordered as in ELEMENT_SET.

    Derived from the WageRateProfile rows actually seeded for the client (data,
    not a fixed list), so a company seeded with a subset of elements gets a
    template with only those columns."""
    codes = {
        p.pay_code
        for p in WageRateProfile.query.filter_by(
            client_company_id=client_company_id
        ).all()
    }
    ordered = [code for code, _l, _c in ELEMENT_SET if code in codes]
    ordered += sorted(codes - set(ordered))  # any non-standard codes, defensively
    return ordered


def generate_monthly_template(client_company_id, export_folder, month, year):
    """Write the monthly thin template for a seeded company and return its path.
    Records an audit entry."""
    employees = (
        Employee.query.filter_by(client_company_id=client_company_id)
        .order_by(Employee.staff_id)
        .all()
    )
    codes = seeded_element_codes(client_company_id)
    header = thin_header(codes) + [ICU_MEMBER_COLUMN]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "MONTHLY TEMPLATE"
    sheet.append(header)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col in range(1, len(header) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for emp in employees:
        row = [emp.staff_id, emp.full_name]
        row += [0] * len(codes)                 # hours, zero-filled
        row += [0, 0, 0, 0, 0]                  # adjustments, zero-filled
        row += ["Member" if emp.icu_member else ""]  # scoped to members only
        sheet.append(row)

    for column_cells in sheet.columns:
        width = max(len(str(c.value or "")) for c in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width + 3, 32)

    from app import db
    from app.models import ClientCompany

    client = db.session.get(ClientCompany, client_company_id)
    client_name = client.name if client else f"Client {client_company_id}"
    filename = (
        f"{slug_filename(client_name)}_Monthly_Template_"
        f"{slug_filename(str(month))}_{year}.xlsx"
    )
    path = save_workbook(workbook, export_folder, filename)

    from app.audit import record_audit

    record_audit(
        "Monthly template generated",
        client,
        f"{client_name} {month} {year}: {len(employees)} workers, "
        f"{len(codes)} element columns.",
    )
    db.session.commit()
    return path
