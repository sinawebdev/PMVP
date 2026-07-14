"""Thin monthly upload: hours + this month's adjustments, joined to seeded
context and run through the Phase 2 compute engine.

The thin file is the system-generated monthly template (Phase 6): one row per
seeded worker, the company's exact hour-element columns, and a small set of
adjustment columns — Prod/Bonus Allowance, Loan, Welfare, Other Deduction and
(optional) Pay Difference. Locked rules (spec §5):

  * **Blank = 0** — never carried forward from a prior month or the seed.
  * **Loan = this month's amount** — no outstanding-balance ledger.
  * **Rates, basic wage, ICU membership and tax relief come from stored
    context**, never the file — so a raise only takes effect after a rich
    re-upload, and ICU is derived (3% of basic for members), never uploaded.
  * **Unknown Staff ID blocks that worker** with a "seed via rich upload first"
    message — new hires and raises are rich uploads only.
"""
from dataclasses import dataclass, field

import openpyxl

from app.models import Employee, WageRateProfile
from app.payroll_calculations import bonus_concession_used_ytd_bulk
from app.raw_engine.calc import apply_rates  # noqa: F401  (kept for parity/readability)
from app.raw_engine.cleaning import coerce_hours, coerce_rate, normalise_element, normalise_emp_id
from app.raw_engine.compute import compute_payslip
from app.raw_engine.mapping import ELEMENT_SET
from app.raw_engine.run import assemble_inputs

# --- thin column contract --------------------------------------------------
STAFF_ID_HEADERS = {"STAFF ID", "STAFF_ID", "STAFFID", "IRS/NO.", "STAFF NO", "ID"}
NAME_HEADERS = {"NAME", "NAMES", "FULL NAME"}
# adjustment role -> accepted header labels (normalised, case-insensitive)
ADJUSTMENT_HEADERS = {
    "bonus": {"PROD/BONUS ALLOWANCE", "PROD / BONUS ALLOWANCE", "PROD BONUS ALLOWANCE",
              "BONUS", "PRODUCTIVITY BONUS", "PROD'TY ALLOW", "PRODUCTIVITY ALLOWANCE"},
    "loan": {"LOAN", "LOAN ADV", "LOAN DEDUCTION", "LOAN ADVANCE"},
    "welfare": {"WELFARE", "WELFARE SUPPLIES"},
    "other_deduction": {"OTHER DEDUCTION", "OTHER DEDUCTIONS"},
    "pay_difference": {"PAY DIFFERENCE", "PAY DIFF", "PAY DIFFERECE"},  # sic tolerated
}
# element display label / pay_code -> canonical pay_code
_ELEMENT_BY_LABEL = {}
for _code, _label, _cat in ELEMENT_SET:
    _ELEMENT_BY_LABEL[normalise_element(_label)] = _code
    _ELEMENT_BY_LABEL[normalise_element(_code)] = _code


_LABEL_BY_CODE = {code: label for code, label, _cat in ELEMENT_SET}
_ADJUSTMENT_COLUMNS = [
    ("bonus", "Prod/Bonus Allowance"),
    ("loan", "Loan"),
    ("welfare", "Welfare"),
    ("other_deduction", "Other Deduction"),
    ("pay_difference", "Pay Difference"),
]


class ThinFormatError(ValueError):
    """The thin upload is not a recognisable monthly template (e.g. no Staff ID
    column) — refuse rather than guess."""


def thin_header(element_codes=None):
    """The canonical thin-template header row: Staff ID, Name, the company's
    element columns (display labels), then the adjustment columns."""
    codes = element_codes or [c for c, _l, _cat in ELEMENT_SET]
    return (
        ["Staff ID", "Name"]
        + [_LABEL_BY_CODE[c] for c in codes]
        + [label for _role, label in _ADJUSTMENT_COLUMNS]
    )


def write_thin_workbook(path, records, element_codes=None):
    """Write a thin monthly workbook (the format Phase 6 generates and this
    module parses). ``records``: iterable of :class:`ThinEmployeeInput` (or
    objects with the same attributes)."""
    codes = element_codes or [c for c, _l, _cat in ELEMENT_SET]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MONTHLY"
    ws.append(thin_header(codes))
    for rec in records:
        row = [rec.staff_id, getattr(rec, "name", "")]
        row += [rec.hours.get(code, 0) for code in codes]
        row += [getattr(rec, role) for role, _label in _ADJUSTMENT_COLUMNS]
        ws.append(row)
    wb.save(path)
    return path


@dataclass
class ThinEmployeeInput:
    staff_id: str
    name: str = ""
    hours: dict = field(default_factory=dict)   # pay_code -> hours
    bonus: float = 0.0
    loan: float = 0.0
    welfare: float = 0.0
    other_deduction: float = 0.0
    pay_difference: float = 0.0


@dataclass
class ThinResult:
    payslips: dict = field(default_factory=dict)   # staff_id -> Payslip
    blocked: list = field(default_factory=list)    # [{staff_id, name, reason}]
    missing_rate_codes: dict = field(default_factory=dict)  # staff_id -> [codes]


def _classify(header_value):
    """Return ('staff_id'|'name'|'element'|'adjustment', detail) or (None, None)."""
    label = normalise_element(header_value)
    if not label:
        return None, None
    if label in STAFF_ID_HEADERS:
        return "staff_id", None
    if label in NAME_HEADERS:
        return "name", None
    if label in _ELEMENT_BY_LABEL:
        return "element", _ELEMENT_BY_LABEL[label]
    for role, names in ADJUSTMENT_HEADERS.items():
        if label in names:
            return "adjustment", role
    return None, None


def _resolve_thin_workbook(source):
    """``(workbook, owns)`` for a path or an already-open Workbook. A path loads
    a light read-only workbook (owns=True → close after); an open Workbook is
    reused (owns=False) so a single upload load serves the thin path too."""
    if isinstance(source, openpyxl.Workbook):
        return source, False
    return openpyxl.load_workbook(source, data_only=True, read_only=True), True


def parse_thin_workbook(source):
    """Parse a thin monthly workbook (a path or an already-open Workbook) into
    ``(list[ThinEmployeeInput], warnings)``. Raises :class:`ThinFormatError` if
    no Staff ID column can be found."""
    wb, owns = _resolve_thin_workbook(source)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        if owns:
            wb.close()
    if not rows:
        raise ThinFormatError("Thin upload is empty.")

    header_idx, colmap = _resolve_header(rows)
    if "staff_id" not in {role for role, _ in colmap.values()}:
        raise ThinFormatError(
            "No 'Staff ID' column found in the thin upload — use the generated "
            "monthly template."
        )

    warnings = []
    inputs = []
    for r in rows[header_idx + 1:]:
        record = _read_row(r, colmap)
        if record is not None:
            inputs.append(record)
    return inputs, warnings


def _resolve_header(rows):
    """Find the header row (first row with a Staff ID column) and map each
    column index to its (role, detail)."""
    for idx, row in enumerate(rows[:8]):
        colmap = {}
        for c, value in enumerate(row):
            role, detail = _classify(value)
            if role:
                colmap[c] = (role, detail)
        if any(role == "staff_id" for role, _ in colmap.values()):
            return idx, colmap
    raise ThinFormatError(
        "No 'Staff ID' column found in the first rows of the thin upload."
    )


def _read_row(row, colmap):
    record = ThinEmployeeInput(staff_id="")
    seen_value = False
    for c, (role, detail) in colmap.items():
        value = row[c] if c < len(row) else None
        if role == "staff_id":
            record.staff_id = normalise_emp_id(value)
        elif role == "name":
            record.name = "" if value is None else str(value).strip()
        elif role == "element":
            hrs = coerce_hours(value)
            if hrs:
                record.hours[detail] = record.hours.get(detail, 0.0) + hrs
                seen_value = True
        elif role == "adjustment":
            amount = coerce_rate(value)
            if amount:
                setattr(record, detail, amount)
                seen_value = True
    if not record.staff_id or record.staff_id == "NAN":
        return None  # blank / spacer / total row
    return record


def join_and_compute(thin_inputs, run, statutory_rate):
    """Join parsed thin inputs to the seeded context for ``run``'s client and
    compute a payslip per worker. Unknown staff IDs are blocked (not costed).
    Returns a :class:`ThinResult`."""
    roster = {
        normalise_emp_id(e.staff_id): e
        for e in Employee.query.filter_by(
            client_company_id=run.client_company_id
        ).all()
    }
    profiles_by_emp, client_defaults = _load_profiles(run.client_company_id)

    used_ytd = bonus_concession_used_ytd_bulk(
        [e.id for e in roster.values()], run.year, exclude_run_id=run.id
    )

    result = ThinResult()
    for rec in thin_inputs:
        emp = roster.get(normalise_emp_id(rec.staff_id))
        if emp is None:
            result.blocked.append({
                "staff_id": rec.staff_id,
                "name": rec.name,
                "reason": (
                    "Unknown Staff ID — no seeded employee. Upload a rich "
                    "(RAW DATA) file to seed this worker first; new hires and "
                    "raises are rich uploads only."
                ),
            })
            continue

        rate_lookup = dict(client_defaults)
        rate_lookup.update(profiles_by_emp.get(emp.id, {}))

        # Explicit classification from the roster (Employee.pay_type); fall back
        # to the rate-table heuristic only for legacy rows seeded before pay_type.
        pay_type = emp.pay_type or (
            "hourly" if any(
                c == WageRateProfile.CATEGORY_BASIC for _r, c in rate_lookup.values()
            ) else "salaried"
        )

        inputs = assemble_inputs(
            rec.staff_id,
            rec.hours,
            rate_lookup,
            pay_type=pay_type,
            employee_id=emp.id,
            basic_fallback=float(emp.basic_salary or 0),
            is_icu_member=bool(emp.icu_member),
            bonus=rec.bonus,
            pay_difference=rec.pay_difference,
            tax_relief_monthly=float(emp.tax_relief_monthly or 0),
            loan=rec.loan,
            welfare=rec.welfare,
            other_deductions=rec.other_deduction,
            bonus_concession_used_ytd=used_ytd.get(emp.id, 0.0),
        )
        if inputs.missing_rate_codes:
            result.missing_rate_codes[rec.staff_id] = list(inputs.missing_rate_codes)
        result.payslips[rec.staff_id] = compute_payslip(inputs, statutory_rate)

    return result


def _load_profiles(client_company_id):
    """Return ``(profiles_by_emp, client_defaults)`` rate lookups:
    ``{employee_id: {pay_code: (rate, category)}}`` and the employee-less
    client-default ``{pay_code: (rate, category)}``."""
    profiles_by_emp = {}
    client_defaults = {}
    for p in WageRateProfile.query.filter_by(
        client_company_id=client_company_id
    ).all():
        entry = (p.hourly_rate, p.category)
        if p.employee_id is None:
            client_defaults[p.pay_code] = entry
        else:
            profiles_by_emp.setdefault(p.employee_id, {})[p.pay_code] = entry
    return profiles_by_emp, client_defaults
