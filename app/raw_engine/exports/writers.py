"""Raw-engine workbook writers — the outputs the standard exporters don't cover.

All reuse the shared `app.excel_utils` primitives (`create_workbook`,
`write_table`, `save_workbook`, `slug_filename`) rather than re-implementing
workbook plumbing.
"""
from openpyxl.styles import Font, PatternFill

from app.excel_utils import (
    create_workbook,
    save_workbook,
    slug_filename,
    write_table,
)
from app.money import money
from app.raw_engine.exports.bank_routing import normalise_bank, route_payments
from app.raw_engine.icu_distribution import distribute_union_dues


def _client_name(payroll_run):
    return payroll_run.client_company.name if payroll_run.client_company else "Client"


def _period(payroll_run):
    return f"{payroll_run.month} {payroll_run.year}"


def export_bank_grouping(payroll_run, export_folder, routing=None, whitelist=None):
    """AKOTO-style bank schedule: banked workers grouped by bank, one block per
    bank with a subtotal, and a grand total equal to the banked net total. This
    is the terminal transfer-submission document."""
    routing = routing or route_payments(payroll_run.items, whitelist=whitelist)
    client_name = _client_name(payroll_run)
    title = f"{client_name} Bank Grouping (AKOTO) {_period(payroll_run)}"
    workbook, sheet = create_workbook(title)

    groups = {}
    for item in routing.banked:
        groups.setdefault(normalise_bank(item.bank_name), []).append(item)

    band_fill = PatternFill("solid", fgColor="052420")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    row = 5
    for bank in sorted(groups):
        items = sorted(groups[bank], key=lambda i: (i.full_name or "").upper())
        band = sheet.cell(row=row, column=1, value=bank)
        band.font = Font(bold=True, color="FFFFFF")
        for col in range(1, 6):
            sheet.cell(row=row, column=col).fill = band_fill
        band.font = Font(bold=True, color="FFFFFF")
        row += 1
        for col, header in enumerate(
            ["Staff ID", "Employee Name", "Bank Branch", "Account Number", "Net Pay (GH¢)"],
            start=1,
        ):
            cell = sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        row += 1
        subtotal = 0.0
        for item in items:
            sheet.cell(row=row, column=1, value=item.staff_id)
            sheet.cell(row=row, column=2, value=item.full_name)
            sheet.cell(row=row, column=3, value=item.bank_branch)
            sheet.cell(row=row, column=4, value=item.bank_account_number)
            sheet.cell(row=row, column=5, value=round(item.net_pay or 0, 2))
            subtotal += item.net_pay or 0
            row += 1
        label = sheet.cell(row=row, column=2, value=f"{bank} Total ({len(items)} workers)")
        label.font = Font(bold=True)
        amount = sheet.cell(row=row, column=5, value=money(subtotal))
        amount.font = Font(bold=True)
        row += 2

    grand = sheet.cell(row=row, column=2, value="GRAND TOTAL (BANK)")
    grand.font = Font(bold=True, size=12)
    grand_amount = sheet.cell(row=row, column=5, value=routing.banked_total)
    grand_amount.font = Font(bold=True, size=12)

    filename = (
        f"{slug_filename(client_name)}_Bank_Grouping_"
        f"{slug_filename(payroll_run.month)}_{payroll_run.year}.xlsx"
    )
    return save_workbook(workbook, export_folder, filename)


def export_pv(payroll_run, export_folder, routing=None, whitelist=None):
    """Payment Voucher (cash) list: every worker not paid by bank — no
    recognised bank or no account number. Grand total equals the PV net total."""
    routing = routing or route_payments(payroll_run.items, whitelist=whitelist)
    client_name = _client_name(payroll_run)
    title = f"{client_name} Payment Voucher (Cash) {_period(payroll_run)}"
    workbook, sheet = create_workbook(title)

    headers = ["Clock/Staff ID", "Employee Name", "Bank on Record", "Net Pay (GH¢)", "Signature"]
    rows = []
    for item in sorted(routing.pv, key=lambda i: (i.full_name or "").upper()):
        rows.append([
            item.staff_id,
            item.full_name,
            item.bank_name or "—",
            round(item.net_pay or 0, 2),
            "",
        ])
    write_table(sheet, 5, headers, rows)
    total_row = 5 + len(rows) + 1
    label = sheet.cell(row=total_row, column=2, value=f"CASH TOTAL ({len(rows)} workers)")
    label.font = Font(bold=True)
    amount = sheet.cell(row=total_row, column=4, value=routing.pv_total)
    amount.font = Font(bold=True)

    filename = (
        f"{slug_filename(client_name)}_Payment_Voucher_"
        f"{slug_filename(payroll_run.month)}_{payroll_run.year}.xlsx"
    )
    return save_workbook(workbook, export_folder, filename)


def export_icu_distribution(payroll_run, export_folder):
    """ICU union-dues remittance: the total ICU deducted, split through the union
    cascade (50/50 → union 75/20/5, edfund 80/20), plus the per-member listing.
    The cascade leaves sum back to the ICU total to the cent."""
    client_name = _client_name(payroll_run)
    title = f"{client_name} ICU Union Distribution {_period(payroll_run)}"
    workbook, sheet = create_workbook(title)

    members = [i for i in payroll_run.items if (i.icu_dues or 0) > 0]
    icu_total = money(sum((i.icu_dues or 0) for i in members))
    dist = distribute_union_dues(icu_total)

    cascade = [
        ["Total ICU dues collected", icu_total],
        ["  Union (50%)", dist.union],
        ["    ICU-Accra (75% of union)", dist.icu_accra],
        ["    Local (20% of union)", dist.local],
        ["    ICU-Tema (5% of union)", dist.icu_tema],
        ["  Edfund (50%)", dist.edfund],
        ["    ICU-EDAC (80% of edfund)", dist.icu_edac],
        ["    DCL-EEF (20% of edfund)", dist.dcl_eef],
        ["Total remitted (leaves)", dist.total_payout],
    ]
    write_table(sheet, 5, ["Distribution", "Amount (GH¢)"], cascade)

    start = 5 + len(cascade) + 2
    sheet.cell(row=start, column=1, value=f"Members ({len(members)})").font = Font(bold=True)
    member_rows = [
        [i.staff_id, i.full_name, round(i.icu_dues or 0, 2)]
        for i in sorted(members, key=lambda i: (i.full_name or "").upper())
    ]
    write_table(sheet, start + 1, ["Staff ID", "Name", "ICU Dues (GH¢)"], member_rows)

    filename = (
        f"{slug_filename(client_name)}_ICU_Distribution_"
        f"{slug_filename(payroll_run.month)}_{payroll_run.year}.xlsx"
    )
    return save_workbook(workbook, export_folder, filename)
