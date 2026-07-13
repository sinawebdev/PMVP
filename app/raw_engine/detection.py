"""Format detection and engine routing for the raw-hours path.

Two questions the upload flow asks before choosing a code path:

  1. Is this workbook a *rich* RAW-DATA seed workbook? (``is_rich_raw_data``)
  2. Has this company already been *seeded*? (``company_is_seeded``)

Routing rule (spec §2): a raw upload for a company with **no** WageRateProfile
rows goes to the seed flow; a company that is already seeded takes the thin
monthly path. Selection is per upload — there is no ``payroll_mode`` column.

**Shape Guard** (``classify_workbook``) is a lightweight, structural pre-check
shared by *both* upload paths so the wrong importer never runs on the wrong
file (the Book1.xlsx case: a Raw Hours workbook fed to the Standard importer
produced 163 invalid rows). It classifies a workbook — before any mapping or
parsing — as one of ``RAW_HOURS`` / ``STANDARD_PAYROLL`` / ``UNKNOWN`` using
only header structure, never employee data. The Standard path blocks a
``RAW_HOURS`` file; the Raw path blocks a ``STANDARD_PAYROLL`` file.
"""
import re

import openpyxl

from app.models import WageRateProfile
from app.raw_engine.mapping import (
    NAME_HEADER_LABEL,
    RAW_DATA_SHEET,
    find_name_header_row,
    HeaderError,
)


def open_raw_data_sheet(path):
    """Load the RAW DATA worksheet (cached values, not formulas). Raises
    HeaderError if the sheet is absent."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    if RAW_DATA_SHEET not in wb.sheetnames:
        raise HeaderError(
            f"Workbook has no '{RAW_DATA_SHEET}' sheet (found: {wb.sheetnames}); "
            "not a DZ-style rich raw workbook."
        )
    return wb[RAW_DATA_SHEET]


def is_rich_raw_data(path) -> bool:
    """True if ``path`` looks like a DZ-style rich RAW-DATA seed workbook: it
    has a RAW DATA sheet with the stacked NAMES header. Never raises — returns
    False on anything it can't confirm."""
    try:
        ws = open_raw_data_sheet(path)
        find_name_header_row(ws)
        return True
    except (HeaderError, Exception):
        return False


def company_is_seeded(client_company_id) -> bool:
    """True if the company already has raw-engine context (any WageRateProfile
    rows). Seeded companies take the thin monthly path; unseeded companies with
    a raw upload go to the seed flow."""
    if not client_company_id:
        return False
    return (
        WageRateProfile.query.filter_by(client_company_id=client_company_id).first()
        is not None
    )


# ── Shape Guard: which importer does this workbook belong to? ─────────────────

# Workbook shape classes returned by ``classify_workbook``.
STANDARD_PAYROLL = "standard_payroll"
RAW_HOURS = "raw_hours"
UNKNOWN = "unknown"

# Hour-category signals that mark a Raw Hours workbook. Each tuple is a set of
# equivalent normalised fragments; a category counts once if ANY fragment is a
# substring of a scanned header cell. Sourced from the raw engine's own element
# labels (``mapping.ELEMENTS`` expected labels + the thin-template display
# labels in ``mapping.ELEMENT_SET``) so detection and parsing can't drift.
# A Standard Payroll sheet carries none of these; a rich RAW DATA sheet or a
# monthly thin template carries all of them.
_RAW_HOUR_CATEGORIES = (
    ("NORMAL",),
    ("WEEK DAY", "WEEKDAY"),
    ("SATURDAY",),
    ("SUN HOL", "SUN HOLIDAY", "SUNHOL"),
    ("AFTERNOON", "AFT NOON"),
    ("NIGHT",),
    ("6 TO 6", "6TO6"),
    ("4CREW", "4 CREW"),
)
# Enough distinct categories to be unambiguous while tolerating a company
# seeded with a subset of elements (a full workbook carries all eight).
_RAW_MIN_CATEGORY_HITS = 3

# Standard Payroll input columns the live resolver must find for a workbook to
# read as a standard sheet — the same bar the importer itself enforces.
_STANDARD_MANDATORY = frozenset({"staff_id", "full_name", "basic_salary"})

# How far down each sheet the structural scan looks (headers sit in the top
# rows; a stacked ACS header is 3 rows, the rich RAW DATA header ~12 rows).
_SCAN_ROWS = 25
_SCAN_COLS = 90


def _norm_token(value) -> str:
    """Upper-case, strip punctuation to single spaces: 'SUN / HOL' -> 'SUN HOL',
    "AFT'NOON" -> 'AFT NOON', 'Normal hours' -> 'NORMAL HOURS'. Lets the raw
    hour-category fragments match a cell regardless of punctuation/case."""
    return re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()


def _scan_header_cells(path):
    """Set of normalised, non-empty cell strings from the top rows of every
    worksheet. Structural only — never reads employee data far down the sheet.
    Returns an empty set on any workbook that cannot be opened (encrypted,
    corrupt, wrong type); callers treat that as 'no signal'."""
    cells = set()
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return cells
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=_SCAN_ROWS, values_only=True):
                for value in row[:_SCAN_COLS]:
                    if value is None:
                        continue
                    token = _norm_token(value)
                    if token:
                        cells.add(token)
    finally:
        wb.close()
    return cells


def looks_like_raw_hours(path) -> bool:
    """True if the workbook carries the Raw Hours hour-category headers (NORMAL,
    WEEK DAY OT, SATURDAY OT, SUN/HOL OT, NIGHT, AFTERNOON, 6-to-6, 4-crew) —
    covering both the rich RAW DATA seed workbook and the monthly thin template.
    Structural and punctuation-tolerant; never raises."""
    cells = _scan_header_cells(path)
    if not cells:
        return False
    hits = 0
    for fragments in _RAW_HOUR_CATEGORIES:
        if any(frag in cell for cell in cells for frag in fragments):
            hits += 1
            if hits >= _RAW_MIN_CATEGORY_HITS:
                return True
    return False


def looks_like_standard_payroll(path) -> bool:
    """True if the live Standard resolver can find the mandatory payroll columns
    (Staff ID, Employee Name, Basic Salary) on some payroll-candidate sheet —
    i.e. the Standard importer would accept this workbook. Reuses
    ``excel_utils`` (the canonical resolver) rather than duplicating mapping
    logic; never raises."""
    # Imported lazily: excel_utils pulls pandas, and the raw engine's own paths
    # must not pay that cost just to import this module.
    from app.excel_utils import extract_payroll_sheet, payroll_sheet_candidates

    try:
        candidates = payroll_sheet_candidates(path)
    except Exception:
        return False
    for candidate in candidates:
        try:
            extraction = extract_payroll_sheet(path, candidate["sheet_name"])
        except Exception:
            continue
        mapped_fields = set(extraction["mapping"].values())
        if _STANDARD_MANDATORY <= mapped_fields:
            return True
    return False


def classify_workbook(path) -> str:
    """Structural Shape-Guard classification, run before any parsing/mapping:

      * ``RAW_HOURS``        — carries the raw hour-category headers.
      * ``STANDARD_PAYROLL`` — the standard resolver finds its mandatory columns.
      * ``UNKNOWN``          — neither signature; let the chosen importer surface
        its own detailed error rather than guessing.

    Raw is tested first: a rich RAW DATA sheet also has a NAMES column and a
    BASIC WAGE column (which would otherwise read as standard), so the
    unambiguous hour-category signal must win. Reuses the same helpers each
    upload path can call directly; never raises."""
    if looks_like_raw_hours(path):
        return RAW_HOURS
    if looks_like_standard_payroll(path):
        return STANDARD_PAYROLL
    return UNKNOWN
