import os
import re
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


logger = logging.getLogger(__name__)


COLUMN_ALIASES = {
    "staff_id": ["staff no", "staff no.", "staff id", "employee id", "emp id", "staff number", "employee no", "worker id", "s/n", "sn", "no", "serial", "irs no"],
    "full_name": ["name", "names", "employee name", "full name", "worker name", "employee", "worker", "officer", "personnel"],
    "client_company": ["client", "client company", "company", "company name", "company assigned", "assigned client"],
    # ID-shaped social-security headers must resolve here, never to the ssnit
    # amount field — hence the explicit "S.S number" / "social security number"
    # variants (and no bare "social security" alias on the amount field below).
    "ssnit_number": ["ssnit no", "ssnit number", "ssnit id", "ssnit contribution number", "s.s number", "ss number", "s.s no", "s s number", "social security number", "social security no"],
    "ghana_card_number": ["ghana card", "ghana card no", "ghana card number", "gh card", "gh card no", "gh card number"],
    "momo_number": ["momo", "momo number", "momo no", "mobile money", "mobile money number", "phone", "phone number", "phone no", "telephone", "tel", "mobile", "mobile number", "mobile no", "cell", "cellphone", "contact number", "contact no"],
    "email": ["email", "e-mail", "e mail", "email address", "email id", "mail"],
    "bank_name": ["bank", "bank name"],
    # "bank branch" must resolve here (exact match wins over the bare "bank"
    # alias above, which would otherwise catch it via substring). Deliberately
    # NO "location"/"bank location" alias: map_columns' substring fallback also
    # matches when the header is a substring of an alias, so a bare "Location"
    # column would map into "bank location" and get mis-filed as a branch.
    "bank_branch": ["bank branch", "branch", "branch name"],
    "bank_account_number": ["account no", "account number", "bank account", "bank account number", "a/c number", "a/c no"],
    "status": ["status", "employee status", "worker status", "employment status"],
    "service_line": ["service line", "department", "unit"],
    "job_role": ["job role", "job title", "role", "position", "designation"],
    "payroll_month": ["payroll month", "month"],
    "basic_salary": ["basic", "basic salary", "base pay", "basic pay", "base salary", "monthly salary", "salary", "basic wage"],
    "transport_allowance": ["transport", "transport allowance", "transportation"],
    "housing_allowance": ["housing", "housing allowance", "rent allowance"],
    "medical_allowance": ["medical", "medical allowance", "medical allow", "med allowance", "med. allowance"],
    # "MEALS" (ACS RAW DATA column L) gets its own column — it used to fold
    # silently into other_allowances, destroying the sheet's audit trail.
    "meal_allowance": ["meal allowance", "meals allowance", "meals", "meal"],
    # end_of_year_bonus is listed BEFORE productivity_bonus so its longer
    # aliases win the substring fallback over productivity's bare "bonus".
    "end_of_year_bonus": ["end of year bonus", "end-of-year bonus", "13th month", "13th month bonus", "annual bonus", "eoy bonus"],
    "productivity_bonus": ["productivity bonus", "productivity", "prod bonus", "prod. bonus", "bonus"],
    "overtime_hours": ["overtime hours", "ot hours"],
    "overtime_pay": ["overtime", "ot pay", "overtime pay", "overtime allowance", "overtime allow"],
    "other_allowances": ["other allowance", "other allowances", "allowances", "allowance"],
    "pay_difference": ["pay difference", "pay diff", "pay differece"],
    "gross_pay": ["gross", "gross pay", "gross salary", "total earnings", "gross earnings", "gross amount"],
    "paye": ["paye", "tax", "income tax", "paye tax", "tax deducted"],
    # No bare "social security" here — that substring caught ID headers like
    # "Social Security Number" and dumped an ID into the amount field.
    "ssnit": ["ssnit", "ssnit contribution", "ssnit employee", "ssnit emp", "ssnit (employee)", "ssnit deduction"],
    "tier_2_pension": ["tier 2", "tier 2 pension", "tier two pension", "pension"],
    "pf_fund_employee": ["pf fund / employee", "pf fund employee", "pf fund", "pf employee", "provident fund", "provident fund employee"],
    # loan_advance before loan_deduction is deliberate NOT: loan_deduction
    # stays first so a bare "Loan" header keeps its historical meaning
    # (a deduction); "loan advance" resolves by exact match.
    "loan_deduction": ["loan deduction", "loan deductions", "loan repayment", "loan", "loan adv"],
    "loan_advance": ["loan advance", "loan advances", "salary advance"],
    # Welfare (ACS column AC) and IOU (AE) get their own columns — previously
    # folded into other_deductions. No bare "iou" alias in the list: the
    # substring fallback would match it inside words like "previous"; a bare
    # "IOU" header still resolves because the normalized header is itself a
    # substring of "iou deduction".
    "welfare_deduction": ["welfare", "welfare deduction", "welfare deductions", "welfare supplies"],
    "iou_deduction": ["iou deduction", "iou deductions", "i o u", "i o u deduction", "iou"],
    "other_deductions": ["deduction", "deductions", "other deduction", "other deductions"],
    "total_deductions": ["total deductions", "total deduction"],
    "net_pay": ["net", "net pay", "net salary", "take home", "take home pay", "net amount", "amount payable", "net earnings"],
}

# Headers for figures the system always derives itself (spec: never read from
# an upload). They must land in unmapped_columns — visible in the preview —
# instead of substring-matching into basic_salary/net_pay and corrupting them.
DERIVED_OUTPUT_HEADERS = {
    "net basic wage",
    "annual salary",
    "15 of annual salary",          # normalize_label("15% of Annual Salary")
    "15 percent of annual salary",
    "15 of annual",
    # Composite labels from the ACS stacked header ("BASIC" over "TAX" ->
    # "BASIC TAX"): every one is a tax figure the engine derives itself. If
    # "basic tax" were allowed through, the substring fallback would bind it
    # to basic_salary — the exact acs 1.xlsx clobber.
    "basic tax",
    "overtime tax",
    "ot tax",
    "o t tax",                      # normalize_label("O.T TAX")
    "bonus tax",
    "total tax",
}

# A payroll import is refused outright when any of these has no column.
MANDATORY_IMPORT_FIELDS = ("staff_id", "full_name", "basic_salary")

# A staff-ID in the first column marks the first data row (AC605, MT035, ...).
# Rows between the detected header row and this anchor are stacked sub-label
# rows (the ACS sheet has label + sub-label + a blank row before data).
STAFF_ID_RE = re.compile(r"^[A-Z]{2,4}\d{2,}$")
STACKED_HEADER_MAX_DEPTH = 5

MONEY_FIELDS = {
    "basic_salary",
    "transport_allowance",
    "housing_allowance",
    "medical_allowance",
    "meal_allowance",
    "productivity_bonus",
    "end_of_year_bonus",
    "pf_fund_employee",
    "overtime_hours",
    "overtime_pay",
    "other_allowances",
    "pay_difference",
    "gross_pay",
    "paye",
    "ssnit",
    "tier_2_pension",
    "loan_deduction",
    "loan_advance",
    "welfare_deduction",
    "iou_deduction",
    "other_deductions",
    "total_deductions",
    "net_pay",
}

META_SHEET_NAMES = {
    "stress test guide",
    "client companies",
    "expected summary",
    "expected validation",
    "upload test cases",
    "summary",
    "guide",
    "instructions",
    "readme",
}

PAYROLL_HEADER_KEYWORDS = {
    "staff id",
    "staff no",
    "staff no.",
    "employee id",
    "emp id",
    "worker id",
    "s/n",
    "sn",
    "serial",
    "employee",
    "name",
    "employee name",
    "full name",
    "worker",
    "officer",
    "personnel",
    "status",
    "service line",
    "job role",
    "basic salary",
    "basic",
    "base pay",
    "basic pay",
    "base salary",
    "monthly salary",
    "gross pay",
    "gross salary",
    "total earnings",
    "gross earnings",
    "gross amount",
    "paye",
    "income tax",
    "tax deducted",
    "ssnit",
    "ssnit employee",
    "ssnit emp",
    "social security",
    "tier 2 pension",
    "net pay",
    "net salary",
    "net amount",
    "amount payable",
    "net earnings",
    "take home",
    "bank",
    "bank name",
    "bank account",
    "momo",
    "ghana card",
    "ghana card no",
    "ssnit number",
    "payroll month",
    "transport allowance",
    "housing allowance",
    "overtime",
    "overtime pay",
    "allowance",
    "deduction",
    "total deductions",
}

PAYROLL_SHEET_NAME_KEYWORDS = {
    "payroll",
    "salary",
    "wages",
    "staff",
    "workers",
    "personnel",
}


def normalize_label(value):
    value = str(value or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_worker(value):
    return normalize_label(value)


def normalize_company_key(value, strip_suffix=True):
    parts = normalize_label(value).split()
    if strip_suffix:
        parts = [part for part in parts if part not in {"ltd", "limited"}]
    return "".join(parts)


def company_tokens(value):
    return {
        part
        for part in normalize_label(value).split()
        if part and part not in {"ltd", "limited", "company", "co"}
    }


def slug_filename(value):
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return cleaned or "report"


def allowed_excel_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"xlsx", "xls", "csv"}


def map_columns(columns):
    mapping = {}
    alias_lookup = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_lookup[normalize_label(alias)] = field

    for column in columns:
        # Duplicate columns arrive suffixed (" (2)", " (3)") from
        # build_composite_columns; map them exactly like their base label so
        # both bind the same field and mapping_conflicts() can refuse the
        # pair — instead of the suffix garbling the lookup (e.g. "ssnit 2"
        # substring-matching staff_id's "sn" alias).
        normalized = normalize_label(re.sub(r"\s*\(\d+\)\s*$", "", str(column)))
        if normalized in DERIVED_OUTPUT_HEADERS:
            # System-derived figure: never accepted from an upload — leave it
            # unmapped so the rep sees it in the preview instead of it
            # silently overwriting basic_salary/net_pay via substring match.
            mapping[column] = "unmapped"
            continue
        # Exact-match only. The former substring fallback (alias in normalized
        # / normalized in alias) let one keyword claim several columns at once
        # — "OVERTIME" binding both OVERTIME ALLOWANCE and EXCESS OVERTIME,
        # "TAX" binding every TAXABLE * column to paye — collisions that
        # mapping_conflicts() then had to refuse. A header matching no alias
        # exactly is left unmapped and surfaced in the preview, never guessed.
        mapping[column] = alias_lookup.get(normalized) or "unmapped"
    return mapping


def mapping_conflicts(mapping, mandatory=MANDATORY_IMPORT_FIELDS):
    """Errors that make a column mapping unsafe to import: a field claimed by
    more than one column (the row reader would silently last-wins — how
    'BASIC' bound basic_salary to the tax column in acs 1.xlsx), or a
    mandatory field with no column at all. Never guesses; the import is
    refused with the specific collision."""
    by_field = {}
    for column, field in mapping.items():
        if field and field != "unmapped":
            by_field.setdefault(field, []).append(str(column))
    errors = []
    for field, columns in by_field.items():
        if len(columns) > 1:
            errors.append(
                f"Field '{field}' is claimed by {len(columns)} columns "
                f"({', '.join(columns)}). Each payroll figure must come from "
                "exactly one column — refusing to guess which."
            )
    for field in mandatory:
        if field not in by_field:
            errors.append(
                f"Mandatory column '{field}' was not found in the sheet "
                "header — the import cannot proceed without it."
            )
    return errors


def find_data_start(rows, header_row, max_depth=STACKED_HEADER_MAX_DEPTH):
    """First data row below the detected header row (0-based), anchored on a
    staff-ID in the first column. ACS-style sheets stack sub-label rows (and
    a blank row) between the label row and the data, so 'header + 1' is
    wrong there. Falls back to header_row + 1 when no anchor is found within
    max_depth rows, or when an intermediate row has a non-empty first cell
    (that's data with a different ID shape, not a sub-label row)."""
    fallback = header_row + 1
    limit = min(len(rows), header_row + 1 + max_depth)
    for index in range(header_row + 1, limit):
        row = rows[index]
        first = str(row[0] if row else "").strip()
        if STAFF_ID_RE.match(first):
            return index
        if first:
            return fallback
    return fallback


def build_composite_columns(rows, header_row, data_start):
    """One label per column from the stacked header block
    (rows[header_row:data_start]): 'BASIC' over 'SALARY' becomes
    'BASIC SALARY', which can never collide with 'BASIC TAX'. Duplicate
    composites get a ' (n)' suffix so two same-named columns stay distinct
    all the way into the mapping, where mapping_conflicts() refuses them
    instead of letting the row reader last-wins."""
    block = rows[header_row:data_start]
    width = max((len(row) for row in block), default=0)
    names, seen = [], {}
    for col in range(width):
        parts = []
        for row in block:
            cell = str(row[col] if col < len(row) else "").strip()
            if cell and cell.lower() not in {"nan", "none"}:
                parts.append(cell)
        name = re.sub(r"\s+", " ", " ".join(parts)).strip()
        if not name:
            name = f"Unnamed: {col}"
        count = seen.get(name, 0) + 1
        seen[name] = count
        if count > 1:
            name = f"{name} ({count})"
        names.append(name)
    return names


def workbook_sheet_names(file_path):
    ext = file_path.rsplit(".", 1)[1].lower()
    if ext == "csv":
        return [Path(file_path).stem]
    with pd.ExcelFile(file_path, engine="openpyxl" if ext == "xlsx" else None) as excel_file:
        return excel_file.sheet_names


def is_meta_sheet(sheet_name):
    normalized = normalize_label(sheet_name)
    return normalized in META_SHEET_NAMES


def sheet_name_suggests_payroll(sheet_name):
    normalized = normalize_label(sheet_name)
    return any(keyword in normalized for keyword in PAYROLL_SHEET_NAME_KEYWORDS)


def known_header_labels():
    """Every normalized string header_score recognises as a column header:
    field names, their aliases, and the loose payroll keywords. Kept identical
    to header_score's original inline set so the misdetection guard that reuses
    it can't perturb header-row scoring."""
    labels = set()
    for field, aliases in COLUMN_ALIASES.items():
        labels.add(normalize_label(field))
        labels.update(normalize_label(alias) for alias in aliases)
    labels.update(PAYROLL_HEADER_KEYWORDS)
    return labels


def looks_like_header_label(value, *, numeric_is_suspicious=False):
    """True when a value that should be human data (a person's name, a company
    name) instead reads like a spreadsheet column header — the signature of the
    header-row detector locking onto the wrong row (as with the "acs 1.xlsx"
    import, where "GH CARD"/"JOB TITLE" header cells landed in the company
    field). Matches the whole cell against any known column alias/label (plus
    the system-derived output headers); when ``numeric_is_suspicious`` is set,
    also flags a bare number like "0" sitting where a name belongs."""
    text = normalize_label(value)
    if not text:
        return False
    if text in known_header_labels() or text in DERIVED_OUTPUT_HEADERS:
        return True
    if numeric_is_suspicious and re.fullmatch(r"-?\d+(?:\.\d+)?", str(value).strip()):
        return True
    return False


def header_score(row):
    known_labels = known_header_labels()

    labels = [normalize_label(value) for value in row if value not in (None, "")]
    score = 0
    for label in labels:
        if label in known_labels:
            score += 2
        elif any(keyword and keyword in label for keyword in PAYROLL_HEADER_KEYWORDS):
            score += 1
    return score


def find_header_row(file_path, sheet_name=None):
    ext = file_path.rsplit(".", 1)[1].lower()
    sheet_arg = sheet_name if sheet_name is not None else 0
    if ext == "csv":
        sample = pd.read_csv(file_path, header=None, nrows=20, dtype=str)
    elif ext == "xlsx":
        sample = pd.read_excel(file_path, sheet_name=sheet_arg, engine="openpyxl", header=None, nrows=20, dtype=str)
    else:
        sample = pd.read_excel(file_path, sheet_name=sheet_arg, header=None, nrows=20, dtype=str)
    rows = sample.fillna("").values.tolist()

    best_row = 1
    best_score = 0
    for row_index, row in enumerate(rows, start=1):
        score = header_score(row)
        if score > best_score:
            best_score = score
            best_row = row_index
    return max(best_row - 1, 0) if best_score >= 2 else 0


def payroll_sheet_candidates(file_path):
    candidates = []
    ext = file_path.rsplit(".", 1)[1].lower()
    for sheet_name in workbook_sheet_names(file_path):
        if is_meta_sheet(sheet_name):
            continue
        header_row = find_header_row(file_path, sheet_name if ext != "csv" else None)
        if ext == "csv":
            sample = pd.read_csv(file_path, header=None, nrows=20, dtype=str)
        elif ext == "xlsx":
            sample = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl", header=None, nrows=20, dtype=str)
        else:
            sample = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20, dtype=str)
        rows = sample.fillna("").values.tolist()
        score = header_score(rows[header_row]) if header_row < len(rows) else 0
        if score >= 2 or sheet_name_suggests_payroll(sheet_name):
            candidates.append(
                {
                    "sheet_name": sheet_name,
                    "detected_header_row": header_row + 1,
                    "score": score,
                }
            )
    return candidates


def match_client_sheet(client_name, sheet_names):
    normalized_client = normalize_company_key(client_name)
    for sheet_name in sheet_names:
        normalized_sheet = normalize_company_key(sheet_name)
        if normalized_client == normalized_sheet:
            return sheet_name
    for sheet_name in sheet_names:
        normalized_sheet = normalize_company_key(sheet_name)
        if normalized_client in normalized_sheet or normalized_sheet in normalized_client:
            return sheet_name
    client_tokens = company_tokens(client_name)
    for sheet_name in sheet_names:
        if len(client_tokens.intersection(company_tokens(sheet_name))) >= 2:
            return sheet_name
    return None


def _read_sheet_frame(file_path, sheet_name=None):
    """DataFrame + mapping + layout metadata for one sheet, resolving stacked
    (multi-row) headers: the header block runs from the scored header row down
    to the staff-ID data anchor, and each column's name is the composite of
    its stacked labels — so 'BASIC SALARY' and 'BASIC TAX' arrive distinct."""
    header_row = find_header_row(file_path, sheet_name)
    ext = file_path.rsplit(".", 1)[1].lower()
    sheet_arg = sheet_name if sheet_name is not None else 0
    if ext == "csv":
        full = pd.read_csv(file_path, header=None, dtype=str)
    elif ext == "xlsx":
        full = pd.read_excel(file_path, sheet_name=sheet_arg, engine="openpyxl", header=None, dtype=str)
    else:
        full = pd.read_excel(file_path, sheet_name=sheet_arg, header=None, dtype=str)
    raw_rows = full.fillna("").values.tolist()
    data_start = find_data_start(raw_rows, header_row)
    columns = build_composite_columns(raw_rows, header_row, data_start)
    df = full.iloc[data_start:].copy()
    df.columns = columns
    df = df.dropna(how="all")
    df = df.fillna("")
    mapping = map_columns(df.columns)
    return {
        "df": df,
        "mapping": mapping,
        "header_row": header_row,
        "data_start": data_start,
    }


def read_excel_file(file_path, sheet_name=None):
    frame = _read_sheet_frame(file_path, sheet_name)
    return frame["df"], frame["mapping"]


def detect_company_name(file_path, known_company_names=None, sheet_name=None):
    known_company_names = known_company_names or []
    sampled_values = []
    ext = file_path.rsplit(".", 1)[1].lower()
    sheet_arg = sheet_name if sheet_name is not None else 0
    if ext == "csv":
        sample = pd.read_csv(file_path, header=None, nrows=20, dtype=str)
    elif ext == "xlsx":
        sample = pd.read_excel(file_path, sheet_name=sheet_arg, engine="openpyxl", header=None, nrows=20, dtype=str)
    else:
        sample = pd.read_excel(file_path, sheet_name=sheet_arg, header=None, nrows=20, dtype=str)
    for row in sample.fillna("").values.tolist():
        for value in row:
            if value not in (None, ""):
                sampled_values.append(str(value).strip())

    combined = " ".join(sampled_values)
    normalized_combined = normalize_label(combined)
    for company_name in known_company_names:
        if normalize_label(company_name) in normalized_combined:
            return company_name

    for label in ("company", "client", "employer"):
        for index, value in enumerate(sampled_values):
            if label in normalize_label(value) and index + 1 < len(sampled_values):
                return sampled_values[index + 1]
    return ""


def to_number(value):
    if hasattr(value, "tolist"):
        values = [item for item in value.tolist() if item not in (None, "")]
        value = values[0] if values else ""
    if value in (None, ""):
        return 0.0
    try:
        cleaned = (
            str(value)
            .replace(",", "")
            .replace("GHS", "")
            .replace("GH₵", "")
            .replace("₵", "")
            .replace("GH¢", "")
            .replace("¢", "")
            .strip()
        )
        cleaned = cleaned.replace("GHC", "")
        cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
        return float(cleaned or 0)
    except (TypeError, ValueError):
        return 0.0


def scalar_cell_value(value):
    if hasattr(value, "tolist"):
        values = [item for item in value.tolist() if item not in (None, "")]
        return values[0] if values else ""
    return value


def mapped_rows_from_dataframe(df, mapping):
    rows = []
    mapped_fields = {field for field in mapping.values() if field != "unmapped"}
    all_unmapped_skips = 0
    for _, source_row in df.iterrows():
        row = {}
        original_presence = {}
        first_cell = scalar_cell_value(source_row.iloc[0]) if len(source_row.index) else ""
        if normalize_label(first_cell) in {"total", "grand total", "subtotal", "summary"}:
            continue
        for original_column, field in mapping.items():
            if field == "unmapped":
                continue
            value = source_row.get(original_column, "")
            value = scalar_cell_value(value)
            original_presence[field] = str(value or "").strip() not in {"", "nan", "None"}
            row[field] = to_number(value) if field in MONEY_FIELDS else str(value).strip()

        if not mapped_fields:
            all_unmapped_skips += 1
            continue

        identity = " ".join(
            str(row.get(field) or "")
            for field in ("staff_id", "full_name", "basic_salary", "gross_pay", "net_pay")
        )
        if not identity.strip() and not any(row.get(field, 0) for field in MONEY_FIELDS):
            continue
        if normalize_label(row.get("staff_id") or row.get("full_name")) in {
            "total",
            "grand total",
            "subtotal",
            "summary",
        }:
            continue

        for field in MONEY_FIELDS:
            row.setdefault(field, 0.0)
        row.setdefault("staff_id", "")
        row.setdefault("full_name", "")
        row.setdefault("ssnit_number", "")
        row.setdefault("ghana_card_number", "")
        row.setdefault("momo_number", "")
        row.setdefault("bank_name", "")
        row.setdefault("bank_branch", "")
        row.setdefault("bank_account_number", "")
        row.setdefault("status", "")
        row.setdefault("service_line", "")
        row.setdefault("job_role", "")
        row.setdefault("payroll_month", "")
        row["_missing_original_net_pay"] = not original_presence.get("net_pay", False)
        row["_missing_original_gross_pay"] = not original_presence.get("gross_pay", False)
        row["_missing_original_total_deductions"] = not original_presence.get("total_deductions", False)

        calculated_gross = (
            row["basic_salary"]
            + row["transport_allowance"]
            + row["housing_allowance"]
            + row["medical_allowance"]
            + row["meal_allowance"]
            + row["productivity_bonus"]
            + row["overtime_pay"]
            + row["other_allowances"]
            + row["pay_difference"]
            + row["end_of_year_bonus"]
        )
        if not row["gross_pay"] and calculated_gross:
            row["gross_pay"] = calculated_gross
        statutory_deductions = row["paye"] + row["ssnit"] + row["tier_2_pension"]
        itemized_other = (
            row["loan_deduction"]
            + row["welfare_deduction"]
            + row["iou_deduction"]
            + row["other_deductions"]
        )
        if row["total_deductions"]:
            pass
        elif row["other_deductions"] >= statutory_deductions and statutory_deductions:
            row["total_deductions"] = row["other_deductions"]
        else:
            row["total_deductions"] = statutory_deductions + itemized_other
        if not row["net_pay"] and row["gross_pay"]:
            row["net_pay"] = row["gross_pay"] - row["total_deductions"]
        rows.append(row)
    logger.debug("Smart Excel Import Engine: rows skipped due to all-unmapped columns=%s", all_unmapped_skips)
    return rows


def calculate_worker_stats(mapped_rows):
    seen = set()
    duplicate_count = 0
    non_blank_rows = 0

    for row in mapped_rows:
        staff_id = normalize_worker(row.get("staff_id"))
        full_name = normalize_worker(row.get("full_name"))
        if not staff_id and not full_name:
            continue

        non_blank_rows += 1
        worker_key = f"staff:{staff_id}" if staff_id else f"name:{full_name}"
        if worker_key in seen:
            duplicate_count += 1
        else:
            seen.add(worker_key)

    return {
        "total_rows": non_blank_rows,
        "total_unique_workers": len(seen),
        "duplicate_count": duplicate_count,
    }


def calculate_status_breakdown(mapped_rows):
    breakdown = {
        "active": 0,
        "inactive": 0,
        "terminated": 0,
        "on_leave": 0,
        "unknown": 0,
    }
    for row in mapped_rows:
        status = normalize_label(row.get("status"))
        if "terminated" in status:
            breakdown["terminated"] += 1
        elif "inactive" in status:
            breakdown["inactive"] += 1
        elif "leave" in status:
            breakdown["on_leave"] += 1
        elif "active" in status:
            breakdown["active"] += 1
        else:
            breakdown["unknown"] += 1
    return breakdown


def summarize_mapped_rows(mapped_rows):
    return {
        "total_rows": len(mapped_rows),
        "valid_rows": len(mapped_rows),
        "invalid_rows": 0,
        "gross_total": sum(float(row.get("gross_pay") or 0) for row in mapped_rows),
        "net_total": sum(float(row.get("net_pay") or 0) for row in mapped_rows),
        "paye_total": sum(float(row.get("paye") or 0) for row in mapped_rows),
        "ssnit_total": sum(float(row.get("ssnit") or 0) for row in mapped_rows),
        "deductions_total": sum(float(row.get("total_deductions") or 0) for row in mapped_rows),
    }


def extract_payroll_sheet(file_path, sheet_name=None):
    frame = _read_sheet_frame(file_path, sheet_name)
    df, mapping = frame["df"], frame["mapping"]
    header_row = frame["header_row"]
    mapping_errors = mapping_conflicts(mapping)
    unmapped_columns = [column for column, field in mapping.items() if field == "unmapped"]
    logger.debug(
        "Smart Excel Import Engine: sheet=%s detected_header_row=%s data_start_row=%s columns=%s mapped_columns=%s unmapped_columns=%s mapping_errors=%s",
        sheet_name or workbook_sheet_names(file_path)[0],
        header_row + 1,
        frame["data_start"] + 1,
        list(df.columns),
        {column: field for column, field in mapping.items() if field != "unmapped"},
        unmapped_columns,
        mapping_errors,
    )
    mapped_rows = mapped_rows_from_dataframe(df, mapping)
    worker_stats = calculate_worker_stats(mapped_rows)
    totals = summarize_mapped_rows(mapped_rows)
    return {
        "sheet_name": sheet_name or workbook_sheet_names(file_path)[0],
        "detected_header_row": header_row + 1,
        "data_start_row": frame["data_start"] + 1,
        "columns": list(df.columns),
        "mapping": mapping,
        "mapping_errors": mapping_errors,
        "preview_rows": df.head(20).astype(str).to_dict(orient="records"),
        "mapped_rows": mapped_rows,
        "worker_stats": worker_stats,
        "status_breakdown": calculate_status_breakdown(mapped_rows),
        "totals": totals,
        "ignored_rows": max(len(df.index) - len(mapped_rows), 0),
    }


# openpyxl raises ValueError if a sheet title contains any of these, regardless
# of length — a client name like "ACS/GMT Shipping" would crash the export.
_INVALID_SHEET_TITLE_CHARS = r'\/?*[]:'


def safe_sheet_title(title):
    """A workbook-safe sheet (tab) title: forbidden characters replaced with a
    space (readable — "ACS/GMT" -> "ACS GMT", not "ACSGMT"), collapsed, and
    truncated to Excel's 31-character limit. Never returns an empty string."""
    cleaned = "".join(
        " " if ch in _INVALID_SHEET_TITLE_CHARS else ch for ch in str(title)
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip()[:31].strip()
    return cleaned or "Sheet"


def create_workbook(report_title):
    workbook = Workbook()
    sheet = workbook.active
    # Only the sheet tab name is constrained by Excel; the A2 header cell keeps
    # the original title (slash and all) so the report still reads correctly.
    sheet.title = safe_sheet_title(report_title)
    sheet["A1"] = "Chrisnat Limited"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = report_title
    sheet["A2"].font = Font(bold=True)
    sheet["A3"] = f"Date generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    return workbook, sheet


def write_table(sheet, start_row, headers, rows):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_index, row in enumerate(rows, start=start_row + 1):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(
            max_length + 3, 35
        )


def save_workbook(workbook, export_folder, filename):
    os.makedirs(export_folder, exist_ok=True)
    file_path = os.path.join(export_folder, filename)
    workbook.save(file_path)
    return file_path


def export_employees(employees, export_folder):
    workbook, sheet = create_workbook("Employee List")
    headers = ["Staff ID", "Name", "Client", "Phone", "SSNIT", "Status", "Basic Salary"]
    rows = [
        [
            employee.staff_id,
            employee.full_name,
            employee.client_company.name if employee.client_company else employee.assigned_client,
            employee.phone,
            employee.ssnit_number,
            employee.status,
            employee.basic_salary,
        ]
        for employee in employees
    ]
    write_table(sheet, 5, headers, rows)
    return save_workbook(workbook, export_folder, "Chrisnat_Employee_List.xlsx")


def export_payroll_run(payroll_run, export_folder):
    client_name = payroll_run.client_company.name if payroll_run.client_company else "Client"
    title = f"{client_name} Payroll {payroll_run.month} {payroll_run.year}"
    workbook, sheet = create_workbook(title)
    headers = [
        "Staff ID",
        "Name",
        "Basic",
        "Transport",
        "Housing",
        "Overtime",
        "Gross",
        "PAYE",
        "SSNIT",
        "Deductions",
        "Net Pay",
        "Warnings",
    ]
    rows = [
        [
            item.staff_id,
            item.full_name,
            item.basic_salary,
            item.transport_allowance,
            item.housing_allowance,
            item.overtime_pay,
            item.gross_pay,
            item.paye,
            item.ssnit,
            item.total_deductions,
            item.net_pay,
            item.warning_notes,
        ]
        for item in payroll_run.items
    ]
    write_table(sheet, 5, headers, rows)
    total_row = len(rows) + 7
    sheet.cell(row=total_row, column=1, value="Totals").font = Font(bold=True)
    sheet.cell(row=total_row, column=7, value=payroll_run.total_gross_pay)
    sheet.cell(row=total_row, column=8, value=payroll_run.total_paye)
    sheet.cell(row=total_row, column=9, value=payroll_run.total_ssnit)
    sheet.cell(row=total_row, column=10, value=payroll_run.total_deductions)
    sheet.cell(row=total_row, column=11, value=payroll_run.total_net_pay)
    filename = (
        f"{slug_filename(client_name)}_Payroll_{payroll_run.month}_{payroll_run.year}.xlsx"
    )
    return save_workbook(workbook, export_folder, filename)


def export_payment_vouchers(vouchers, export_folder):
    workbook, sheet = create_workbook("Payment Vouchers")
    headers = [
        "Voucher",
        "Client",
        "Month",
        "Workers",
        "Gross Payroll",
        "Deductions",
        "Net Payable",
        "Status",
        "Prepared By",
        "Reviewed By",
        "Approved By",
    ]
    rows = [
        [
            voucher.voucher_number,
            voucher.payroll_run.client_company.name if voucher.payroll_run.client_company else "",
            f"{voucher.payroll_run.month} {voucher.payroll_run.year}",
            voucher.payroll_run.total_workers,
            voucher.gross_payroll,
            voucher.total_deductions,
            voucher.net_amount_payable,
            voucher.status,
            voucher.preparer.name if voucher.preparer else "",
            voucher.reviewer.name if voucher.reviewer else "",
            voucher.approver.name if voucher.approver else "",
        ]
        for voucher in vouchers
    ]
    write_table(sheet, 5, headers, rows)
    return save_workbook(workbook, export_folder, "Chrisnat_Payment_Vouchers.xlsx")


def export_remittances(remittances, export_folder):
    workbook, sheet = create_workbook("Remittance Summary")
    headers = ["Client", "Month", "Type", "Amount Due", "Due Date", "Status", "Reference"]
    rows = [
        [
            remittance.payroll_run.client_company.name if remittance.payroll_run.client_company else "",
            f"{remittance.payroll_run.month} {remittance.payroll_run.year}",
            remittance.remittance_type,
            remittance.amount_due,
            remittance.due_date.isoformat() if remittance.due_date else "",
            remittance.status,
            remittance.payment_reference,
        ]
        for remittance in remittances
    ]
    write_table(sheet, 5, headers, rows)
    return save_workbook(workbook, export_folder, "Chrisnat_Remittance_Summary.xlsx")


def export_expenses(expenses, export_folder):
    workbook, sheet = create_workbook("Expense List")
    headers = ["Date", "Category", "Description", "Amount", "Method", "Receipt"]
    rows = [
        [
            expense.expense_date.isoformat(),
            expense.category,
            expense.description,
            expense.amount,
            expense.payment_method,
            expense.receipt_reference,
        ]
        for expense in expenses
    ]
    write_table(sheet, 5, headers, rows)
    total_row = len(rows) + 7
    sheet.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
    sheet.cell(row=total_row, column=4, value=sum(expense.amount for expense in expenses))
    return save_workbook(workbook, export_folder, "Chrisnat_Expenses.xlsx")


def export_monthly_payroll_summary(payroll_runs, export_folder, month, year):
    workbook, sheet = create_workbook(f"Monthly payroll summary {month} {year}")
    headers = [
        "Client",
        "Month",
        "Workers",
        "Gross Payroll",
        "Deductions",
        "PAYE",
        "SSNIT",
        "Net Payroll",
        "Status",
    ]
    rows = [
        [
            run.client_company.name if run.client_company else "",
            f"{run.month} {run.year}",
            run.total_workers,
            run.total_gross_pay,
            run.total_deductions,
            run.total_paye,
            run.total_ssnit,
            run.total_net_pay,
            run.status,
        ]
        for run in payroll_runs
    ]
    write_table(sheet, 5, headers, rows)
    total_row = len(rows) + 7
    sheet.cell(row=total_row, column=1, value="Totals").font = Font(bold=True)
    sheet.cell(row=total_row, column=3, value=sum(run.total_workers for run in payroll_runs))
    sheet.cell(row=total_row, column=4, value=sum(run.total_gross_pay for run in payroll_runs))
    sheet.cell(row=total_row, column=5, value=sum(run.total_deductions for run in payroll_runs))
    sheet.cell(row=total_row, column=6, value=sum(run.total_paye for run in payroll_runs))
    sheet.cell(row=total_row, column=7, value=sum(run.total_ssnit for run in payroll_runs))
    sheet.cell(row=total_row, column=8, value=sum(run.total_net_pay for run in payroll_runs))
    return save_workbook(
        workbook,
        export_folder,
        f"Chrisnat_Monthly_Payroll_{slug_filename(month)}_{year}.xlsx",
    )


def export_import_error_report(payload, export_folder):
    workbook, sheet = create_workbook("Payroll Import Error Report")
    headers = ["Row", "Staff ID", "Name", "Bank Account", "Net Pay", "Warnings"]
    rows = []
    for row_number, warnings in payload.get("validation", {}).get("per_row_warnings", {}).items():
        index = int(row_number) - 1
        mapped_row = payload.get("mapped_rows", [])[index] if index < len(payload.get("mapped_rows", [])) else {}
        rows.append(
            [
                row_number,
                mapped_row.get("staff_id", ""),
                mapped_row.get("full_name", ""),
                mapped_row.get("bank_account_number", ""),
                mapped_row.get("net_pay", 0),
                "; ".join(warnings),
            ]
        )
    write_table(sheet, 5, headers, rows)
    return save_workbook(
        workbook,
        export_folder,
        f"Import_Errors_{slug_filename(payload.get('source_filename', 'payroll'))}.xlsx",
    )


def export_bank_listing(payroll_run, export_folder):
    """Bank transfer batch listing: employees grouped by bank, each row showing
    account number and net pay, with a subtotal per bank. Derived entirely from
    the run's items — replaces the per-bank tables maintained by hand in the
    source workbooks."""
    client_name = payroll_run.client_company.name if payroll_run.client_company else "Client"
    title = f"{client_name} Bank Listing {payroll_run.month} {payroll_run.year}"
    workbook, sheet = create_workbook(title)

    groups = {}
    for item in payroll_run.items:
        bank = (item.bank_name or "").strip() or "NO BANK ON RECORD"
        groups.setdefault(bank, []).append(item)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    band_fill = PatternFill("solid", fgColor="052420")
    row_index = 5
    for bank in sorted(groups):
        items = sorted(groups[bank], key=lambda i: (i.full_name or "").upper())
        band = sheet.cell(row=row_index, column=1, value=bank)
        band.font = Font(bold=True, color="FFFFFF")
        band.fill = band_fill
        for col in range(2, 6):
            sheet.cell(row=row_index, column=col).fill = band_fill
        row_index += 1
        for col_index, header in enumerate(
            ["Staff ID", "Employee Name", "Bank Branch", "Account Number", "Net Pay (GH¢)"], start=1
        ):
            cell = sheet.cell(row=row_index, column=col_index, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        row_index += 1
        subtotal = 0.0
        for item in items:
            sheet.cell(row=row_index, column=1, value=item.staff_id)
            sheet.cell(row=row_index, column=2, value=item.full_name)
            sheet.cell(row=row_index, column=3, value=item.bank_branch)
            sheet.cell(row=row_index, column=4, value=item.bank_account_number)
            sheet.cell(row=row_index, column=5, value=round(item.net_pay or 0, 2))
            subtotal += item.net_pay or 0
            row_index += 1
        total_cell = sheet.cell(row=row_index, column=2, value=f"{bank} Total ({len(items)} workers)")
        total_cell.font = Font(bold=True)
        amount_cell = sheet.cell(row=row_index, column=5, value=round(subtotal, 2))
        amount_cell.font = Font(bold=True)
        row_index += 2

    grand = sheet.cell(row=row_index, column=2, value="GRAND TOTAL")
    grand.font = Font(bold=True, size=12)
    grand_amount = sheet.cell(
        row=row_index, column=5,
        value=round(sum(i.net_pay or 0 for i in payroll_run.items), 2),
    )
    grand_amount.font = Font(bold=True, size=12)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 3, 40)
    filename = (
        f"{slug_filename(client_name)}_Bank_Listing_"
        f"{slug_filename(payroll_run.month)}_{payroll_run.year}.xlsx"
    )
    return save_workbook(workbook, export_folder, filename)


def export_wages_sheet(payroll_run, export_folder):
    """Wages Sheet export matching Chrisnat's own ACS "WAGE SHT" tab: 17
    columns in the client's exact order, plus a totals row over every money
    column. All figures come from the run's items — the SSF/derived columns
    are calculator output persisted at Calculate/confirm time."""
    client_name = payroll_run.client_company.name if payroll_run.client_company else "Client"
    title = f"{client_name} Wages Sheet {payroll_run.month} {payroll_run.year}"
    workbook, sheet = create_workbook(title)
    headers = [
        "Staff ID",
        "Name",
        "Basic Wage",
        "SSF 5.5%",
        "SSF 13%",
        "Net Basic Wage",
        "Transport Allowance",
        "Pay Difference",
        "Other Allowance",
        "Overtime Allowance",
        "Gross Pay",
        "Provident Fund",
        "Other Deductions",
        "Tax (PAYE)",
        "Net Pay",
    ]
    rows = []
    for item in sorted(payroll_run.items, key=lambda i: (i.full_name or "").upper()):
        rows.append([
            item.staff_id,
            item.full_name,
            round(item.basic_salary or 0, 2),
            round(item.ssnit or 0, 2),
            round(item.ssf_employer or 0, 2),
            round(item.net_basic_wage or 0, 2),
            round(item.transport_allowance or 0, 2),
            round(item.pay_difference or 0, 2),
            round(item.other_allowances or 0, 2),
            round(item.overtime_pay or 0, 2),
            round(item.gross_pay or 0, 2),
            round(item.pf_fund_employee or 0, 2),
            # WAGE SHT "OTHER DEDUCTIONS" is the fold of every non-statutory
            # deduction: loan + welfare + other + IOU (spec §9) — the sheet
            # has no separate columns for them, and the row only reconciles
            # to net pay if all four are in here.
            round(
                (item.loan_deduction or 0)
                + (item.welfare_deduction or 0)
                + (item.other_deductions or 0)
                + (item.iou_deduction or 0),
                2,
            ),
            round(item.paye or 0, 2),
            round(item.net_pay or 0, 2),
        ])
    write_table(sheet, 5, headers, rows)
    total_row = 5 + len(rows) + 1
    total_label = sheet.cell(row=total_row, column=2, value="TOTALS")
    total_label.font = Font(bold=True)
    for column in range(3, len(headers) + 1):
        total = round(sum(row[column - 1] or 0 for row in rows), 2)
        cell = sheet.cell(row=total_row, column=column, value=total)
        cell.font = Font(bold=True)
    filename = (
        f"{slug_filename(client_name)}_Wages_Sheet_"
        f"{slug_filename(payroll_run.month)}_{payroll_run.year}.xlsx"
    )
    return save_workbook(workbook, export_folder, filename)


GRA_TAX_OFFICES = ("LTO", "MTO", "STO")


def format_tax_office_tickboxes(tax_office):
    """Render the GRA form's LTO/MTO/STO tick-boxes from CHRISNAT_TAX_OFFICE.
    An unrecognised value is appended verbatim rather than dropped."""
    selected = str(tax_office or "").strip().upper()
    boxes = "   ".join(
        f"{office} [{'X' if office == selected else ' '}]"
        for office in GRA_TAX_OFFICES
    )
    if selected and selected not in GRA_TAX_OFFICES:
        boxes += f"   Other: {tax_office}"
    return boxes


# The official GRA "PAYE DATA 2022" workbook, shipped as a skeleton so the
# export is byte-faithful to the statutory form. We overwrite its embedded
# 2022 tax formulas with the app's StatutoryRate-computed VALUES: keeping the
# 2022 bands would print tax that disagrees with the payslips we already issue.
GRA_TEMPLATE_PATH = Path(__file__).resolve().parent / "assets" / "gra_paye_template.xlsx"

# Data rows begin at row 18 on the GRA template (rows 1-17 are the header block
# and the column-number/label rows).
_GRA_DATA_START_ROW = 18

_MONTH_NUMBERS = {
    name: number
    for number, name in enumerate(
        [
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December",
        ],
        start=1,
    )
}


def _gra_month_code(month, year):
    """MM/YYYY for the 'FOR THE MONTH OF' cell; falls back to the raw month
    string if it isn't a recognised English month name."""
    key = str(month or "").strip().title()
    number = _MONTH_NUMBERS.get(key)
    return f"{number:02d}/{year}" if number else f"{month} {year}"


def _fill_gra_skeleton_workbook(payroll_run, employer_tin, tax_office):
    """Load the official GRA workbook and pour this run's already-computed
    figures into cells A-AB, one worker per row from row 18. GRA's own formula
    cells (I, M, O, S, U, V, W, Y, Z) are overwritten with the app's values, so
    the filed schedule reconciles to the payslips exactly. Columns with no
    backing data (accommodation/vehicle/non-cash elements, secondary employment,
    severance) are left blank for hand-filling."""
    from openpyxl import load_workbook

    def m(value):
        return round(float(value or 0), 2)

    workbook = load_workbook(GRA_TEMPLATE_PATH)
    sheet = workbook.active

    # --- header block (write to the top-left cell of each merged range) ---
    sheet["D9"] = "CHRISNAT LIMITED"                       # NAME OF EMPLOYER
    sheet["D11"] = employer_tin or ""                      # EMPLOYER'S TIN
    sheet["U9"] = _gra_month_code(payroll_run.month, payroll_run.year)
    if tax_office:
        sheet["G7"] = tax_office                           # Name of Tax Office

    count = 0
    for offset, item in enumerate(
        sorted(payroll_run.items, key=lambda i: (i.full_name or "").upper())
    ):
        row = _GRA_DATA_START_ROW + offset
        count += 1

        total_bonus = (item.productivity_bonus or 0) + (item.end_of_year_bonus or 0)
        bonus_concession = m(max(total_bonus - (item.bonus_excess or 0), 0))
        cash_allowances = m(
            (item.transport_allowance or 0)
            + (item.housing_allowance or 0)
            + (item.medical_allowance or 0)
            + (item.meal_allowance or 0)
            + (item.other_allowances or 0)
            + (item.pay_difference or 0)
        )
        # Marginal-band tax only: strip the concessionary overtime/bonus tax
        # back out of the persisted total (col W is "Tax Deductible" on chargeable
        # income; overtime and bonus tax get their own columns Y and M).
        ordinary_paye = m((item.paye or 0) - (item.overtime_tax or 0) - (item.bonus_tax or 0))
        employee = getattr(item, "employee", None)
        # Total Reliefs (U) = SSF + Third Tier only. Deductible Reliefs (T) is
        # left blank: the payroll system has no reliable relief data, and GRA
        # columns without data stay on the sheet but empty (never deleted).
        total_reliefs = m((item.ssnit or 0) + (item.pf_fund_employee or 0))
        tin = (getattr(employee, "tin", None) if employee else None) or item.ghana_card_number or ""

        cells = {
            "A": count,
            "B": tin,
            "C": item.full_name,
            "D": item.job_role,
            "E": "",                                 # Residency — no reliable data; blank
            "F": m(item.basic_salary),
            "G": "",                                 # Secondary Employment — no data; blank
            "H": "Y" if (item.ssnit or 0) > 0 else "N",   # Paid SSNIT
            "I": m(item.ssnit),                      # Social Security Fund
            "J": m(item.pf_fund_employee),           # Third Tier
            "K": cash_allowances,                    # Cash Allowances
            "L": bonus_concession,                   # Bonus up to 15% annual basic
            "M": m(item.bonus_tax),                  # Final Tax on Bonus Income
            "N": m(item.bonus_excess),               # Excess Bonus
            "O": m(item.gross_pay),                  # Total Cash Emolument
            "P": "", "Q": "", "R": "",               # accommodation/vehicle/non-cash
            "S": m(item.gross_pay),                  # Total Assessable Income
            "T": "",                                 # Deductible Reliefs — no data; blank
            "U": total_reliefs,                      # Total Reliefs (SSF + Third Tier)
            "V": m(item.taxable_income),             # Chargeable Income
            "W": ordinary_paye,                      # Tax Deductible
            "X": m(item.overtime_pay),               # Overtime Income
            "Y": m(item.overtime_tax),               # Overtime Tax
            "Z": m(item.paye),                       # Total Tax Payable to GRA
            "AA": "",                                # Severance Pay
            "AB": item.warning_notes or "",          # Remarks
        }
        for column, value in cells.items():
            sheet[f"{column}{row}"] = value          # overwrites GRA's 2022 formula

    # Drop the template's unused formula rows below the last worker so the file
    # doesn't ship hundreds of blank =IF(...) rows.
    first_unused = _GRA_DATA_START_ROW + count
    if sheet.max_row >= first_unused:
        sheet.delete_rows(first_unused, sheet.max_row - first_unused + 1)

    return workbook


def export_gra_paye_schedule(payroll_run, export_folder, employer_tin="", tax_office=""):
    """Employer's Monthly Tax Deductions Schedule (P.A.Y.E.) in the statutory
    GRA format. Primary path loads the official GRA template (byte-faithful form)
    and fills computed VALUES into A-AB. If the template asset is missing, it
    falls back to the summary layout below so exports never hard-fail."""
    if GRA_TEMPLATE_PATH.exists():
        client_name = (
            payroll_run.client_company.name if payroll_run.client_company else "Client"
        )
        workbook = _fill_gra_skeleton_workbook(payroll_run, employer_tin, tax_office)
        filename = (
            f"{slug_filename(client_name)}_GRA_PAYE_Schedule_"
            f"{slug_filename(payroll_run.month)}_{payroll_run.year}.xlsx"
        )
        return save_workbook(workbook, export_folder, filename)
    return _export_gra_summary_layout(
        payroll_run, export_folder, employer_tin, tax_office
    )


def _export_gra_summary_layout(payroll_run, export_folder, employer_tin="", tax_office=""):
    """Fallback summary layout (used only when the GRA template asset is absent).
    The employer of record is always CHRISNAT LIMITED — Chrisnat is the legal
    employer regardless of the client site a worker is deployed to; the client
    name appears only as deployment context. Columns with no backing data yet
    (TIN where unset, Non-Resident, Secondary Employment, benefit elements,
    Severance, Remark) are left blank for hand-filling in Excel."""
    client_name = payroll_run.client_company.name if payroll_run.client_company else "Client"
    workbook, sheet = create_workbook(
        f"GRA PAYE {payroll_run.month} {payroll_run.year}"
    )
    sheet["A2"] = "EMPLOYER'S MONTHLY TAX DEDUCTIONS SCHEDULE (P.A.Y.E.)"
    sheet["A2"].font = Font(bold=True)
    sheet["A4"] = "Name of Employer: CHRISNAT LIMITED"
    sheet["A4"].font = Font(bold=True)
    sheet["A5"] = f"Employer TIN: {employer_tin or ''}"
    sheet["A6"] = f"Tax Office (tick one): {format_tax_office_tickboxes(tax_office)}"
    sheet["A7"] = f"Client / Deployment Site: {client_name}"
    sheet["A8"] = f"Month: {payroll_run.month} {payroll_run.year}"

    headers = [
        "No.",
        "Employee Name",
        "Staff ID",
        "TIN",
        "Ghana Card No.",
        "SSNIT No.",
        "Non-Resident (Y/N)",
        "Secondary Employment (Y/N)",
        "Basic Salary",
        "Total Cash Emoluments",
        "Employee SSF",
        "Third Tier / Provident Fund",
        "Accommodation Element",
        "Vehicle Element",
        "Non-Cash Benefit",
        "Chargeable Income",
        "Tax Deducted (PAYE)",
        "Overtime Income",
        "Overtime Tax",
        "Bonus Income",
        "Final Tax on Bonus Income",
        "Excess Bonus",
        "Severance Pay",
        "Total Tax Payable to GRA",
        "Remark",
    ]
    rows = []
    for index, item in enumerate(
        sorted(payroll_run.items, key=lambda i: (i.full_name or "").upper()), start=1
    ):
        # Concession bonus income = total bonus minus the excess that joined
        # ordinary taxable income (both persisted by the calculator).
        total_bonus = (item.productivity_bonus or 0) + (item.end_of_year_bonus or 0)
        bonus_concession = round(max(total_bonus - (item.bonus_excess or 0), 0), 2)
        ordinary_paye = round(
            (item.paye or 0) - (item.overtime_tax or 0) - (item.bonus_tax or 0), 2
        )
        rows.append([
            index,
            item.full_name,
            item.staff_id,
            item.employee.tin if item.employee else "",
            item.ghana_card_number,
            item.ssnit_number,
            "",  # Non-Resident — no backing data, hand-fill if applicable
            "",  # Secondary Employment — same
            round(item.basic_salary or 0, 2),
            round(item.gross_pay or 0, 2),
            round(item.ssnit or 0, 2),
            round(item.pf_fund_employee or 0, 2),
            "",  # Accommodation Element
            "",  # Vehicle Element
            "",  # Non-Cash Benefit
            round(item.taxable_income or 0, 2),
            ordinary_paye,
            round(item.overtime_pay or 0, 2),
            round(item.overtime_tax or 0, 2),
            bonus_concession,
            round(item.bonus_tax or 0, 2),
            round(item.bonus_excess or 0, 2),
            "",  # Severance Pay
            round(item.paye or 0, 2),
            "",  # Remark
        ])
    write_table(sheet, 10, headers, rows)
    total_row = 10 + len(rows) + 1
    sheet.cell(row=total_row, column=2, value="TOTALS").font = Font(bold=True)
    money_columns = (9, 10, 11, 12, 16, 17, 18, 19, 20, 21, 22, 24)
    for column in money_columns:
        total = round(sum(row[column - 1] or 0 for row in rows), 2)
        cell = sheet.cell(row=total_row, column=column, value=total)
        cell.font = Font(bold=True)
    filename = (
        f"{slug_filename(client_name)}_GRA_PAYE_Schedule_"
        f"{slug_filename(payroll_run.month)}_{payroll_run.year}.xlsx"
    )
    return save_workbook(workbook, export_folder, filename)
