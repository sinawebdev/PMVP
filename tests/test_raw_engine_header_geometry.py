"""RAW DATA header-geometry tolerance for the seed layout validator.

Regression cover for the two real header shapes the DZ RAW DATA workbook ships
in, plus a fail-loud guard. The original ``validate_layout`` hardcoded
``element_row = name_row - 2`` with exact-label matching, verified against a
single specimen; a client workbook that collapsed the stacked header into a
merged two-row header (NAMES one row higher, category labels merged onto the
NAMES row, hour labels suffixed like 'NORMAL HOURS') was rejected with
'RATE 1 != NORMAL' because the parser read the rate-tier row.

These fixtures are built in-memory (no client PII), so the case runs everywhere
— unlike ``test_raw_engine_seed`` which needs the decrypted DZ specimen.
"""
import os
import unittest

os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")

from openpyxl import Workbook

from app.raw_engine.mapping import (
    HeaderError,
    find_element_row,
    find_name_header_row,
    validate_layout,
)

# 1-based columns mirrored from mapping.py so the fixtures place labels exactly
# where the validator reads them.
HOURS_COLS = range(4, 13)   # D..L
RATE_COLS = range(14, 23)   # N..V
COL_DAILY_RATE, COL_BASIC_WAGE, COL_ICU_DUES = 13, 23, 55

# Suffixed hour labels (the merged Book1 variant) and the bare rate labels that
# sit on the same element row in both variants.
HOURS_LABELS = [
    "NORMAL HOURS", "WEEK DAY OT", "SATURDAY OT", "SUN / HOL OT",
    "AFT'NOON ALLOWANCE", "NIGHT ALLOWANCE", "6 TO 6 DAY", "6 TO 6 NIGHT",
    "4crew HOURS",
]
BARE_LABELS = [
    "NORMAL", "WEEK DAY", "SATURDAY", "SUN / HOL", "AFT'NOON", "NIGHT",
    "6 TO 6", "6 TO 6", "4crew",
]
SUB_LABELS = ["HOURS", "OT", "OT", "OT", "ALLOW.", "ALLOW.", "DAY", "NIGHT", "HOURS"]


def _blank_raw_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "RAW DATA"
    return ws


def _fill_rate_and_anchors(ws):
    """Rate block (N..V) + daily-rate anchor (M) — identical in both variants,
    all on the element row (10)."""
    ws.cell(10, COL_DAILY_RATE, "RATE")
    for col, lab in zip(RATE_COLS, BARE_LABELS):
        ws.cell(10, col, lab)


def build_merged_two_row():
    """Book1 geometry: category labels merged vertically D10:D11 (value in the
    anchor, row 11 reads blank), NAMES on row 11, W/BC merged 'BASIC WAGE' /
    'ICU DUES', data from row 12. name_row == 11."""
    ws = _blank_raw_sheet()
    for col, lab in zip(HOURS_COLS, HOURS_LABELS):
        ws.cell(10, col, lab)
        ws.merge_cells(start_row=10, start_column=col, end_row=11, end_column=col)
    _fill_rate_and_anchors(ws)
    ws.cell(10, COL_BASIC_WAGE, "BASIC WAGE")
    ws.merge_cells(start_row=10, start_column=COL_BASIC_WAGE, end_row=11, end_column=COL_BASIC_WAGE)
    ws.cell(10, COL_ICU_DUES, "ICU DUES")
    ws.merge_cells(start_row=10, start_column=COL_ICU_DUES, end_row=11, end_column=COL_ICU_DUES)
    ws.cell(11, 2, "NAMES")
    ws.cell(12, 1, 1)
    ws.cell(12, 2, "GEORGE AKOTO")
    return ws


def build_stacked_three_row():
    """DZ geometry: bare category labels on row 10, a sub-label row 11
    (HOURS/OT/ALLOW), NAMES on row 12, W/BC split as 'BASIC'/'WAGE' and
    'ICU'/'DUES' across rows 10/11, data from row 13. name_row == 12."""
    ws = _blank_raw_sheet()
    for col, lab, sub in zip(HOURS_COLS, BARE_LABELS, SUB_LABELS):
        ws.cell(10, col, lab)
        ws.cell(11, col, sub)
    _fill_rate_and_anchors(ws)
    ws.cell(10, COL_BASIC_WAGE, "BASIC")
    ws.cell(11, COL_BASIC_WAGE, "WAGE")
    ws.cell(10, COL_ICU_DUES, "ICU")
    ws.cell(11, COL_ICU_DUES, "DUES")
    ws.cell(12, 2, "NAMES")
    ws.cell(13, 1, 1)
    ws.cell(13, 2, "GEORGE AKOTO")
    return ws


class HeaderGeometryTests(unittest.TestCase):
    def test_merged_two_row_validates(self):
        ws = build_merged_two_row()
        name_row = find_name_header_row(ws)
        self.assertEqual(name_row, 11)
        self.assertEqual(find_element_row(ws, name_row), 10)
        validate_layout(ws, name_row)  # must not raise

    def test_stacked_three_row_validates(self):
        ws = build_stacked_three_row()
        name_row = find_name_header_row(ws)
        self.assertEqual(name_row, 12)
        self.assertEqual(find_element_row(ws, name_row), 10)
        validate_layout(ws, name_row)  # must not raise

    def test_both_geometries_resolve_the_same_element_row(self):
        """Merged and stacked headers both anchor the element row on NAMES, so
        the seed reads identical data columns regardless of header shape."""
        merged = build_merged_two_row()
        stacked = build_stacked_three_row()
        self.assertEqual(
            find_element_row(merged, find_name_header_row(merged)),
            find_element_row(stacked, find_name_header_row(stacked)),
        )

    def test_garbage_header_still_fails_loud(self):
        """A NAMES row with unrelated headers is refused (fail-loud contract),
        not silently accepted and not crashed on a non-positive row index."""
        ws = _blank_raw_sheet()
        ws.cell(1, 2, "NAMES")
        for col in range(3, 23):
            ws.cell(1, col, f"junk{col}")
        ws.cell(2, 2, "SOME PERSON")
        with self.assertRaises(HeaderError):
            validate_layout(ws, find_name_header_row(ws))


if __name__ == "__main__":
    unittest.main()
