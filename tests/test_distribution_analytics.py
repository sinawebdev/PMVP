"""Phase 4, Slice 3 — delivery analytics + export reports.

Analytics aggregates the filtered delivery set by channel and company; exports
stream it as CSV / XLSX. Operator-plane, reusing the history filters.
"""

import os
import unittest

os.environ["SKIP_DOTENV"] = "true"
os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["SEED_DEMO_DATA"] = "true"
os.environ["PERSISTENCE_REQUIRED"] = "false"

from app import create_app, db  # noqa: E402
from app.distribution.analytics import (  # noqa: E402
    delivery_analytics,
    export_deliveries_csv,
    export_deliveries_xlsx,
)
from app.distribution.queue import enqueue_distribution, process_all_queued  # noqa: E402
from app.models import PayrollRun, User  # noqa: E402


class AnalyticsTestCase(unittest.TestCase):
    def setUp(self):
        self.app = create_app()
        self.app.config["TESTING"] = True
        self.ctx = self.app.app_context()
        self.ctx.push()
        self.run = PayrollRun.query.filter_by(status="Approved").first()
        self.operator = User.query.filter_by(email="admin@chrisnat.local").first()
        # One item fails (no contact) so analytics has a sent/failed mix on sms.
        item = self.run.items[0]
        item.momo_number = None
        item.email = None
        if item.employee:
            item.employee.phone = None
            item.employee.momo_number = None
            item.employee.email = None
        db.session.commit()
        enqueue_distribution(self.run, "sms", False, self.operator)
        process_all_queued()

    def tearDown(self):
        db.session.remove()
        self.ctx.pop()

    def test_analytics_group_by_channel_and_company(self):
        stats = delivery_analytics({})
        self.assertGreater(stats["totals"]["total"], 0)
        self.assertEqual(
            stats["totals"]["total"],
            stats["totals"]["sent"] + stats["totals"]["failed"],
        )
        sms = [r for r in stats["by_channel"] if r["key"] == "sms"]
        self.assertTrue(sms)
        self.assertGreaterEqual(sms[0]["failed"], 1)
        # Company breakdown carries a resolved name.
        self.assertTrue(all("name" in r for r in stats["by_company"]))
        self.assertGreaterEqual(
            stats["totals"]["success_rate"] + stats["totals"]["failure_rate"], 99.9
        )

    def test_status_filter_narrows_analytics(self):
        sent_only = delivery_analytics({"status": "sent"})
        self.assertEqual(sent_only["totals"]["failed"], 0)
        self.assertGreater(sent_only["totals"]["sent"], 0)

    def test_csv_export_has_header_and_rows(self):
        data, filename = export_deliveries_csv({})
        text = data.decode("utf-8-sig")
        self.assertTrue(filename.endswith(".csv"))
        self.assertIn("Staff ID", text.splitlines()[0])
        self.assertGreater(len(text.splitlines()), 1)

    def test_xlsx_export_is_a_real_workbook(self):
        data, filename = export_deliveries_xlsx({})
        self.assertTrue(filename.endswith(".xlsx"))
        self.assertTrue(data.startswith(b"PK"))  # zip/xlsx magic
        # Round-trips through openpyxl with a Summary sheet.
        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data))
        self.assertIn("Deliveries", wb.sheetnames)
        self.assertIn("Summary", wb.sheetnames)


class AnalyticsRouteTestCase(unittest.TestCase):
    def setUp(self):
        self.app = create_app()
        self.app.config["TESTING"] = True
        self.http = self.app.test_client()
        self.ctx = self.app.app_context()
        self.ctx.push()
        self.run = PayrollRun.query.filter_by(status="Approved").first()
        self.operator = User.query.filter_by(email="admin@chrisnat.local").first()
        enqueue_distribution(self.run, "auto", False, self.operator)
        process_all_queued()

    def tearDown(self):
        db.session.remove()
        self.ctx.pop()

    def _login(self, email):
        self.http.post("/login", data={"email": email, "password": "password123"})

    def test_operator_sees_analytics_and_exports(self):
        self._login("admin@chrisnat.local")
        self.assertEqual(self.http.get("/distribution/analytics").status_code, 200)
        csv_resp = self.http.get("/distribution/history/export.csv")
        self.assertEqual(csv_resp.status_code, 200)
        self.assertIn("text/csv", csv_resp.mimetype)
        self.assertIn("attachment", csv_resp.headers["Content-Disposition"])
        xlsx_resp = self.http.get("/distribution/history/export.xlsx")
        self.assertEqual(xlsx_resp.status_code, 200)
        self.assertTrue(xlsx_resp.get_data().startswith(b"PK"))

    def test_tenant_user_is_blocked(self):
        self._login("admin@msc.demo")
        resp = self.http.get("/distribution/analytics", follow_redirects=False)
        self.assertEqual(resp.status_code, 302)
        self.assertNotIn("/distribution/analytics", resp.headers["Location"])


if __name__ == "__main__":
    unittest.main()
