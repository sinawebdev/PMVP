"""Tests for the bank-branch field, the delete-button template fix, and the
defensive header-misdetection warning.

Covers three changes:
  1. Delete button renders for every DELETABLE_STATUS (Draft/Previewed/Rejected)
     and stays hidden for others — it used to be hardcoded to "Draft" only.
  2. bank_branch round-trips import -> roster -> grid edit -> bank listing export.
  3. A misdetected header row now surfaces a visible warning instead of silently
     persisting garbage (the "acs 1.xlsx" signature).
"""
import os
import tempfile
import unittest

os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["SEED_DEMO_DATA"] = "true"

from openpyxl import Workbook, load_workbook

from app import create_app, db
from app.excel_utils import (
    export_bank_listing,
    looks_like_header_label,
    map_columns,
    mapped_rows_from_dataframe,
    read_excel_file,
    detect_company_name,
)
from app.models import ClientCompany, Employee, PayrollItem, PayrollRun
from app.payroll import (
    DELETABLE_STATUSES,
    create_or_update_employee_from_import,
)
from app.payroll_status import APPROVED, DRAFT, PROCESSED, REJECTED
from app.validators import validate_payroll_rows, validate_single_row


class DeleteButtonRenderingTestCase(unittest.TestCase):
    """Item 1: the delete button follows DELETABLE_STATUSES, not a hardcoded
    'Draft'."""

    def setUp(self):
        self.app = create_app()
        self.app.config["TESTING"] = True
        self.http = self.app.test_client()
        self.http.post(
            "/login", data={"email": "admin@chrisnat.local", "password": "password123"}
        )
        with self.app.app_context():
            self.client_id = ClientCompany.query.first().id

    def _make_run(self, status):
        with self.app.app_context():
            run = PayrollRun(
                client_company_id=self.client_id, month="August", year=2026,
                status=status,
            )
            db.session.add(run)
            db.session.commit()
            return run.id

    def _delete_form_present(self, run_id):
        resp = self.http.get(f"/payroll/runs/{run_id}")
        self.assertEqual(resp.status_code, 200)
        return f"/runs/{run_id}/delete".encode() in resp.data

    def test_button_renders_for_every_deletable_status(self):
        # Exactly the backend's own set — no second hardcoded copy to drift.
        for status in DELETABLE_STATUSES:
            with self.subTest(status=status):
                run_id = self._make_run(status)
                self.assertTrue(
                    self._delete_form_present(run_id),
                    f"delete button should render for {status!r}",
                )

    def test_button_hidden_for_non_deletable_statuses(self):
        for status in (APPROVED, PROCESSED, "Pending Approval"):
            with self.subTest(status=status):
                run_id = self._make_run(status)
                self.assertFalse(
                    self._delete_form_present(run_id),
                    f"delete button must NOT render for {status!r}",
                )

    def test_rejected_is_deletable(self):
        self.assertIn(REJECTED, DELETABLE_STATUSES)


class LooksLikeHeaderLabelTestCase(unittest.TestCase):
    """Item 3: the misdetection guard recognises column headings and placeholders."""

    def test_flags_known_column_headings(self):
        # Whole-cell matches against known aliases/labels — the exact strings
        # that landed in data fields when acs 1.xlsx misparsed.
        for label in ("GH CARD", "JOB TITLE", "COMPANY ASSIGNED",
                      "BASIC SALARY", "ACCOUNT NUMBER", "BANK BRANCH"):
            with self.subTest(label=label):
                self.assertTrue(looks_like_header_label(label))

    def test_flags_bare_number_only_when_asked(self):
        self.assertFalse(looks_like_header_label("0"))
        self.assertTrue(looks_like_header_label("0", numeric_is_suspicious=True))
        self.assertTrue(looks_like_header_label("12", numeric_is_suspicious=True))

    def test_does_not_flag_real_names_or_companies(self):
        for real in ("SAMPSON K. KLUVIE", "David Kwame Tetteh", "ACS/GMT Shipping",
                     "MSC Ghana Ltd"):
            with self.subTest(real=real):
                self.assertFalse(looks_like_header_label(real, numeric_is_suspicious=True))


class BankBranchColumnMappingTestCase(unittest.TestCase):
    """Item 2: 'bank branch' resolves to bank_branch without stealing 'bank'."""

    def test_branch_headers_map_to_bank_branch(self):
        mapping = map_columns(["Bank", "Bank Branch", "Branch", "Account Number"])
        self.assertEqual(mapping["Bank"], "bank_name")
        self.assertEqual(mapping["Bank Branch"], "bank_branch")
        self.assertEqual(mapping["Branch"], "bank_branch")
        self.assertEqual(mapping["Account Number"], "bank_account_number")

    def test_plain_location_does_not_map_to_branch(self):
        # Deliberately excluded alias: a generic work-location column must not
        # be swallowed by bank_branch.
        mapping = map_columns(["Location"])
        self.assertNotEqual(mapping.get("Location"), "bank_branch")


class BankBranchRoundTripTestCase(unittest.TestCase):
    """Item 2: branch survives import -> roster -> grid edit -> bank listing."""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.app = create_app()
        self.app.config["TESTING"] = True
        self.http = self.app.test_client()
        self.http.post(
            "/login", data={"email": "admin@chrisnat.local", "password": "password123"}
        )
        self.ctx = self.app.app_context()
        self.ctx.push()
        self.client_co = ClientCompany.query.first()

    def tearDown(self):
        self.ctx.pop()
        self.temp_dir.cleanup()

    def _write_import_file(self):
        path = os.path.join(self.temp_dir.name, "with_branch.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Staff No", "Name", "Basic Salary", "Bank", "Bank Branch",
                   "Account Number", "Net Pay"])
        ws.append(["AC900", "KOFI TEST", 1000, "ACCESS", "TEMA C1", "0123", 1000])
        wb.save(path)
        return path

    def test_import_populates_employee_bank_branch(self):
        # 1) IMPORT: the smart engine must map + carry bank_branch.
        path = self._write_import_file()
        df, mapping = read_excel_file(path)
        rows = mapped_rows_from_dataframe(df, mapping)
        self.assertEqual(rows[0]["bank_branch"], "TEMA C1")
        self.assertEqual(rows[0]["bank_name"], "ACCESS")

        run = PayrollRun(client_company_id=self.client_co.id, month="July",
                         year=2026, status=DRAFT)
        db.session.add(run)
        db.session.flush()
        emp = create_or_update_employee_from_import(
            rows[0], self.client_co, run, 0, {}
        )
        db.session.commit()
        # 2) ROSTER: the branch reflects on the employee record.
        self.assertEqual(emp.bank_branch, "TEMA C1")
        self.assertEqual(emp.bank_name, "ACCESS")

    def test_grid_edit_updates_item_bank_branch(self):
        run = PayrollRun(client_company_id=self.client_co.id, month="July",
                         year=2026, status=DRAFT)
        db.session.add(run)
        db.session.flush()
        item = PayrollItem(payroll_run_id=run.id, staff_id="AC900",
                           full_name="KOFI TEST", bank_name="ACCESS",
                           bank_branch="OLD BRANCH", net_pay=1000)
        db.session.add(item)
        db.session.commit()
        run_id, item_id = run.id, item.id

        # 3) GRID EDIT: change the branch in the raw-figures grid.
        resp = self.http.post(
            f"/payroll/runs/{run_id}/items/edit",
            data={f"item-{item_id}-bank_branch": "TEMA C1"},
            follow_redirects=True,
        )
        self.assertEqual(resp.status_code, 200)
        db.session.expire_all()
        self.assertEqual(db.session.get(PayrollItem, item_id).bank_branch, "TEMA C1")

    def test_bank_listing_export_has_branch_column(self):
        run = PayrollRun(client_company_id=self.client_co.id, month="July",
                         year=2026, status=APPROVED, total_net_pay=1000)
        db.session.add(run)
        db.session.flush()
        db.session.add(PayrollItem(
            payroll_run_id=run.id, staff_id="AC900", full_name="KOFI TEST",
            bank_name="ACCESS", bank_branch="TEMA C1",
            bank_account_number="0123", net_pay=1000,
        ))
        db.session.commit()

        # 4) BANK LISTING EXPORT: branch column present and correct.
        path = export_bank_listing(run, self.temp_dir.name)
        wb = load_workbook(path)
        cells = {(c.row, c.column): c.value for row in wb.active.iter_rows() for c in row}
        values = [v for v in cells.values() if v is not None]
        self.assertIn("Bank Branch", values)
        self.assertIn("TEMA C1", values)


class HeaderMisdetectionWarningTestCase(unittest.TestCase):
    """Item 3: a misdetected header row surfaces a visible warning."""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.app = create_app()
        self.app.config["TESTING"] = True
        self.ctx = self.app.app_context()
        self.ctx.push()
        self.client_co = ClientCompany.query.first()

    def tearDown(self):
        self.ctx.pop()
        self.temp_dir.cleanup()

    def test_header_label_company_name_no_longer_warns(self):
        # Company detection is retired (PMVP Investigation 02): the company is
        # authoritative from the selected client, so a header-like string is
        # never surfaced as a company warning. Guards against reintroducing the
        # noisy "GH CARD" mismatch/heading warnings.
        result = validate_payroll_rows(
            [{"staff_id": "AC1", "full_name": "KOFI TEST", "net_pay": 100,
              "paye": 10, "ssnit": 5}],
            self.client_co, "July", 2026,
            detected_company_name="GH CARD",
        )
        self.assertFalse(
            any("looks like a spreadsheet column heading" in w
                or "appears to mention" in w
                for w in result["summary_warnings"]),
            result["summary_warnings"],
        )

    def test_name_that_is_a_placeholder_warns(self):
        # "0" is the acs 1.xlsx signature; "JOB TITLE" is a header label landing
        # in the name column.
        for bad_name in ("0", "JOB TITLE"):
            with self.subTest(name=bad_name):
                warnings = validate_single_row({"staff_id": "AC1", "full_name": bad_name})
                self.assertTrue(
                    any("looks like a column heading or placeholder" in w
                        for w in warnings),
                    warnings,
                )

    def test_real_name_does_not_warn(self):
        warnings = validate_single_row({"staff_id": "AC1", "full_name": "KOFI TEST"})
        self.assertFalse(
            any("looks like a column heading" in w for w in warnings), warnings
        )

    def test_synthetic_shifted_file_end_to_end(self):
        # A workbook whose real header row sits below junk, with a company
        # marker adjacent to a header cell and a "0" where a name belongs —
        # the acs 1.xlsx failure shape, built from scratch.
        path = os.path.join(self.temp_dir.name, "shifted.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["MONTHLY PAYROLL REPORT"])
        ws.append([])
        ws.append(["IRS/NO.", "NAMES", "BASIC SALARY", "JOB TITLE",
                   "COMPANY ASSIGNED", "GH CARD", "ACCOUNT NUMBER", "NET PAY"])
        ws.append(["AC1", "0", 1000, "RELEASE", "ACS", "GHA-1", "012", 1000])
        wb.save(path)

        detected = detect_company_name(path, ["MSC Ghana Ltd"])
        self.assertTrue(
            looks_like_header_label(detected),
            f"detected company {detected!r} should look like a header label",
        )

        df, mapping = read_excel_file(path)
        rows = mapped_rows_from_dataframe(df, mapping)
        self.assertTrue(rows, "expected at least one data row")
        row_warnings = validate_single_row(rows[0])
        self.assertTrue(
            any("looks like a column heading or placeholder" in w
                for w in row_warnings),
            row_warnings,
        )


if __name__ == "__main__":
    unittest.main()
