import os
import tempfile
import unittest
from datetime import date, datetime
from io import BytesIO

os.environ["DATABASE_URL"] = "sqlite:///:memory:"
os.environ["SEED_DEMO_DATA"] = "true"

from openpyxl import Workbook

import app as app_module
from app import create_app, format_ghana_cedis
from app.pdf_service import generate_payslip_pdf
from app.excel_utils import calculate_worker_stats, map_columns
from app import db
from app.models import AuditTrail, ClientCompany, Employee, Expense, PayrollItem, PayrollRun, User


class MvpTestCase(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.app = create_app()
        self.app.config["TESTING"] = True
        self.app.config["EXPORT_FOLDER"] = self.temp_dir.name
        self.client = self.app.test_client()

    def tearDown(self):
        self.temp_dir.cleanup()

    def login_admin(self):
        return self.client.post(
            "/login",
            data={"email": "admin@chrisnat.local", "password": "password123"},
            follow_redirects=True,
        )

    def login_md(self):
        return self.client.post(
            "/login",
            data={"email": "md@chrisnat.local", "password": "password123"},
            follow_redirects=True,
        )

    def login_operations(self):
        return self.client.post(
            "/login",
            data={"email": "operations@chrisnat.local", "password": "password123"},
            follow_redirects=True,
        )

    def login_client(self, email="msc.client@chrisnat.local"):
        return self.client.post(
            "/login",
            data={"email": email, "password": "password123"},
            follow_redirects=True,
        )

    def build_import_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "staff id",
                "full name",
                "ssnit no",
                "basic salary",
                "transport allowance",
                "housing allowance",
                "overtime pay",
                "gross pay",
                "paye",
                "ssnit",
                "other deductions",
                "net pay",
            ]
        )
        sheet.append(["NEW-101", "New Import Worker", "SSNIT-NEW-101", 2100, 100, 100, 0, 2300, 100, 80, 0, 2120])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def build_offset_phase2_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Chrisnat Limited", "Payroll schedule"])
        sheet.append(["Client", "MSC Ghana Ltd"])
        sheet.append([])
        sheet.append(
            [
                "Employee No",
                "Employee Name",
                "Bank",
                "Account Number",
                "Basic Salary",
                "Allowances",
                "Gross Salary",
                "PAYE Tax",
                "SSNIT Contribution",
                "Deductions",
                "Net Salary",
            ]
        )
        sheet.append(["P2-001", "Akua Boateng", "GCB Bank", "0012345678", "GHC 1,000.00", "200", "", "50", "30", "20", "1,100.00"])
        sheet.append(["P2-001", "Akua Boateng", "GCB Bank", "0012345678", "GHC 1,000.00", "200", "", "50", "30", "20", "1,100.00"])
        sheet.append(["", "", "", "", "", "", "", "", "", "", ""])
        sheet.append(["TOTAL", "", "", "", "", "", "", "", "", "", ""])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def build_multisheet_stress_workbook(self):
        workbook = Workbook()
        guide = workbook.active
        guide.title = "Stress_Test_Guide"
        guide.append(["This workbook starts with a guide sheet and should not be imported."])
        guide.append(["The importer must inspect the payroll sheets that follow."])

        sheet = workbook.create_sheet("MSC_Ghana_Ltd")
        sheet.append(["Chrisnat Limited", "Client payroll schedule"])
        sheet.append(["Client Company", "MSC Ghana Ltd"])
        sheet.append(["Prepared for stress testing"])
        sheet.append(
            [
                "Row ID",
                "Staff ID",
                "Employee Name",
                "Status",
                "Service Line",
                "Job Role",
                "Payroll Month",
                "Basic Salary",
                "Transport Allowance",
                "Housing Allowance",
                "Overtime Pay",
                "Other Allowances",
                "Gross Pay",
                "PAYE",
                "SSNIT Employee",
                "Tier 2 Pension",
                "Loan Deduction",
                "Other Deduction",
                "Total Deductions",
                "Net Pay",
                "Bank Name",
                "Bank Account",
                "MoMo Number",
                "Ghana Card No",
                "SSNIT Number",
            ]
        )
        sheet.append(
            [
                1,
                "MSC-001",
                "Abena Owusu",
                "Active",
                "Port Services",
                "Clerk",
                "May 2026",
                "GHC 1,200.00",
                "150",
                "100",
                "50",
                "25",
                "1,525.00",
                "120",
                "80",
                "40",
                "0",
                "10",
                "250",
                "1,275.00",
                "GCB Bank",
                "0012345678",
                "0244000001",
                "GHA-111",
                "SSNIT-111",
            ]
        )
        sheet.append(
            [
                2,
                "MSC-001",
                "Abena Owusu",
                "Inactive",
                "Port Services",
                "Clerk",
                "May 2026",
                "GHC 1,200.00",
                "150",
                "100",
                "50",
                "25",
                "1,525.00",
                "120",
                "80",
                "40",
                "0",
                "10",
                "250",
                "1,275.00",
                "GCB Bank",
                "0012345678",
                "0244000001",
                "GHA-111",
                "SSNIT-111",
            ]
        )
        sheet.append(
            [
                3,
                "",
                "Kojo Mensah",
                "Terminated",
                "Terminal",
                "Operator",
                "May 2026",
                "900",
                "100",
                "",
                "",
                "",
                "",
                "0",
                "0",
                "",
                "",
                "",
                "0",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        sheet.append(["TOTAL", "", "", "", "", "", "", "", "", "", "", "", "3,950.00", "", "", "", "", "", "", ""])

        stellar = workbook.create_sheet("Stellar_Logistics")
        stellar.append(["Metadata"])
        stellar.append([])
        stellar.append(["Staff ID", "Employee Name", "Gross Pay", "Net Pay"])
        stellar.append(["STL-001", "Afia Darko", 1000, 900])

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def build_multiclient_workbook(self):
        workbook = Workbook()
        guide = workbook.active
        guide.title = "README"
        guide.append(["Guide sheet, not payroll"])

        for sheet_name, staff_id, worker_name, gross, net in [
            ("MSC_Ghana_Ltd", "MSC-MC-001", "Adwoa Frimpong", 1500, 1300),
            ("Stellar_Logistics", "STL-MC-001", "Yaw Antwi", 1000, 900),
        ]:
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(["Client payroll export"])
            sheet.append([])
            sheet.append(["Staff ID", "Employee Name", "Status", "Gross Pay", "PAYE", "SSNIT", "Net Pay"])
            sheet.append([staff_id, worker_name, "Active", gross, 100, 50, net])

        unmatched = workbook.create_sheet("Unknown_Client_Payroll")
        unmatched.append(["Staff ID", "Employee Name", "Gross Pay", "Net Pay"])
        unmatched.append(["UNK-001", "Unmatched Worker", 800, 700])

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def build_consolidated_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Consolidated_Mixed_Clients"
        sheet.append(["Chrisnat Limited", "Mixed client payroll"])
        sheet.append([])
        sheet.append(["Client Company", "Staff ID", "Worker", "Gross Amount", "Tax Deducted", "SSNIT Emp", "Net Amount"])
        sheet.append(["MSC Ghana Ltd", "MSC-C-001", "Efua Mensah", "GHC 1,200.00", "100", "50", "1,050.00"])
        sheet.append(["Stellar Logistics", "STL-C-001", "Kofi Adu", "900", "80", "40", "780"])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def build_minimal_payroll_named_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "May Staff Wages"
        sheet.append(["s/n", "officer", "monthly salary", "gross earnings", "income tax", "ssnit (employee)", "amount payable"])
        sheet.append(["001", "Minimal Worker", "1000", "1200", "90", "55", "1055"])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def test_seeded_users_and_clients_exist(self):
        with self.app.app_context():
            self.assertIsNotNone(User.query.filter_by(email="admin@chrisnat.local").first())
            self.assertIsNotNone(ClientCompany.query.filter_by(name="MSC Ghana Ltd").first())

    def test_database_url_normalizes_render_postgres_url(self):
        previous_url = os.environ.get("DATABASE_URL")
        os.environ["DATABASE_URL"] = "postgres://user:pass@example.com/dbname"
        try:
            self.assertEqual(
                app_module.resolve_database_uri("/tmp/local.db"),
                "postgresql://user:pass@example.com/dbname",
            )
        finally:
            if previous_url is None:
                os.environ.pop("DATABASE_URL", None)
            else:
                os.environ["DATABASE_URL"] = previous_url

    def test_production_requires_persistent_database_url(self):
        previous_url = os.environ.get("DATABASE_URL")
        previous_render = os.environ.get("RENDER")
        previous_flask_env = os.environ.get("FLASK_ENV")
        previous_skip_dotenv = os.environ.get("SKIP_DOTENV")
        os.environ.pop("DATABASE_URL", None)
        os.environ["RENDER"] = "true"
        os.environ["SKIP_DOTENV"] = "true"
        os.environ.pop("FLASK_ENV", None)
        try:
            with self.assertRaises(RuntimeError) as context:
                create_app()
            self.assertIn("DATABASE_URL", str(context.exception))
        finally:
            if previous_url is None:
                os.environ.pop("DATABASE_URL", None)
            else:
                os.environ["DATABASE_URL"] = previous_url
            if previous_render is None:
                os.environ.pop("RENDER", None)
            else:
                os.environ["RENDER"] = previous_render
            if previous_flask_env is None:
                os.environ.pop("FLASK_ENV", None)
            else:
                os.environ["FLASK_ENV"] = previous_flask_env
            if previous_skip_dotenv is None:
                os.environ.pop("SKIP_DOTENV", None)
            else:
                os.environ["SKIP_DOTENV"] = previous_skip_dotenv

    def test_default_seed_does_not_create_demo_payroll_without_flag(self):
        previous_url = os.environ.get("DATABASE_URL")
        previous_seed = os.environ.get("SEED_DEMO_DATA")
        sqlite_path = os.path.join(self.temp_dir.name, "starter_only.db")
        os.environ["DATABASE_URL"] = f"sqlite:///{sqlite_path}"
        os.environ["SEED_DEMO_DATA"] = "false"
        try:
            starter_app = create_app()
            with starter_app.app_context():
                self.assertGreater(User.query.count(), 0)
                self.assertGreater(ClientCompany.query.count(), 0)
                self.assertEqual(PayrollRun.query.count(), 0)
                db.session.remove()
                db.engine.dispose()
        finally:
            if previous_url is None:
                os.environ.pop("DATABASE_URL", None)
            else:
                os.environ["DATABASE_URL"] = previous_url
            if previous_seed is None:
                os.environ.pop("SEED_DEMO_DATA", None)
            else:
                os.environ["SEED_DEMO_DATA"] = previous_seed

    def test_payroll_records_persist_across_app_restart_with_same_database(self):
        previous_url = os.environ.get("DATABASE_URL")
        previous_seed = os.environ.get("SEED_DEMO_DATA")
        previous_persistence = os.environ.get("PERSISTENCE_REQUIRED")
        sqlite_path = os.path.join(self.temp_dir.name, "persistent_payroll.db")
        os.environ["DATABASE_URL"] = f"sqlite:///{sqlite_path}"
        os.environ["SEED_DEMO_DATA"] = "false"
        os.environ["PERSISTENCE_REQUIRED"] = "false"
        try:
            first_app = create_app()
            with first_app.app_context():
                admin = User.query.filter_by(email="admin@chrisnat.local").first()
                client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
                payroll_run = PayrollRun(
                    month="December",
                    year=2099,
                    status="Draft",
                    created_by=admin.id,
                    client_company_id=client_company.id,
                    total_workers=1,
                    total_net_pay=1234.56,
                    source_filename="restart-proof.xlsx",
                )
                db.session.add(payroll_run)
                db.session.commit()
                db.session.remove()
                db.engine.dispose()

            second_app = create_app()
            with second_app.app_context():
                persisted_run = PayrollRun.query.filter_by(
                    month="December",
                    year=2099,
                    source_filename="restart-proof.xlsx",
                ).first()
                self.assertIsNotNone(persisted_run)
                self.assertEqual(persisted_run.total_net_pay, 1234.56)
                db.session.remove()
                db.engine.dispose()
        finally:
            if previous_url is None:
                os.environ.pop("DATABASE_URL", None)
            else:
                os.environ["DATABASE_URL"] = previous_url
            if previous_seed is None:
                os.environ.pop("SEED_DEMO_DATA", None)
            else:
                os.environ["SEED_DEMO_DATA"] = previous_seed
            if previous_persistence is None:
                os.environ.pop("PERSISTENCE_REQUIRED", None)
            else:
                os.environ["PERSISTENCE_REQUIRED"] = previous_persistence

    def test_column_mapping_handles_common_payroll_headers(self):
        mapping = map_columns(["Staff No", "Employee Name", "Basic Salary", "Take Home"])

        self.assertEqual(mapping["Staff No"], "staff_id")
        self.assertEqual(mapping["Employee Name"], "full_name")
        self.assertEqual(mapping["Basic Salary"], "basic_salary")
        self.assertEqual(mapping["Take Home"], "net_pay")

    def test_phase2_import_detects_offset_headers_and_cleans_payroll_values(self):
        from app.excel_utils import mapped_rows_from_dataframe, read_excel_file

        workbook = self.build_offset_phase2_workbook()
        file_path = os.path.join(self.temp_dir.name, "offset_payroll.xlsx")
        with open(file_path, "wb") as file:
            file.write(workbook.read())

        df, mapping = read_excel_file(file_path)
        rows = mapped_rows_from_dataframe(df, mapping)

        self.assertEqual(mapping["Employee No"], "staff_id")
        self.assertEqual(mapping["Employee Name"], "full_name")
        self.assertEqual(mapping["Gross Salary"], "gross_pay")
        self.assertEqual(mapping["Net Salary"], "net_pay")
        self.assertEqual(mapping["Account Number"], "bank_account_number")
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["staff_id"], "P2-001")
        self.assertEqual(rows[0]["bank_account_number"], "0012345678")
        self.assertEqual(rows[0]["gross_pay"], 1200.0)
        self.assertEqual(rows[0]["net_pay"], 1100.0)

    def test_import_handles_duplicate_excel_headers_without_crashing(self):
        import pandas as pd
        from app.excel_utils import mapped_rows_from_dataframe

        df = pd.DataFrame(
            [["CN-900", "Duplicate Header Worker", 1200, 95, 1105]],
            columns=["Staff ID", "Employee Name", "Gross Pay", "Gross Pay", "Net Pay"],
        )
        mapping = {
            "Staff ID": "staff_id",
            "Employee Name": "full_name",
            "Gross Pay": "gross_pay",
            "Net Pay": "net_pay",
        }

        rows = mapped_rows_from_dataframe(df, mapping)

        self.assertEqual(rows[0]["gross_pay"], 1200.0)
        self.assertEqual(rows[0]["net_pay"], 1105.0)

    def test_multisheet_import_matches_selected_client_sheet_and_offset_header(self):
        from app.excel_utils import extract_payroll_sheet, match_client_sheet, payroll_sheet_candidates

        workbook = self.build_multisheet_stress_workbook()
        file_path = os.path.join(self.temp_dir.name, "stress_test.xlsx")
        with open(file_path, "wb") as file:
            file.write(workbook.read())

        candidates = payroll_sheet_candidates(file_path)
        matched_sheet = match_client_sheet("MSC Ghana Ltd", [candidate["sheet_name"] for candidate in candidates])
        extraction = extract_payroll_sheet(file_path, matched_sheet)

        self.assertEqual(matched_sheet, "MSC_Ghana_Ltd")
        self.assertEqual(extraction["detected_header_row"], 4)
        self.assertEqual(extraction["worker_stats"]["total_rows"], 3)
        self.assertEqual(extraction["worker_stats"]["total_unique_workers"], 2)
        self.assertEqual(extraction["worker_stats"]["duplicate_count"], 1)
        self.assertEqual(extraction["status_breakdown"]["active"], 1)
        self.assertEqual(extraction["status_breakdown"]["inactive"], 1)
        self.assertEqual(extraction["status_breakdown"]["terminated"], 1)
        self.assertEqual(extraction["totals"]["gross_total"], 4050.0)
        self.assertEqual(extraction["totals"]["paye_total"], 240.0)
        self.assertEqual(extraction["totals"]["ssnit_total"], 160.0)
        self.assertEqual(extraction["mapping"]["SSNIT Employee"], "ssnit")
        self.assertEqual(extraction["mapping"]["Bank Account"], "bank_account_number")

    def test_minimal_payroll_sheet_name_and_aliases_extract_rows(self):
        from app.excel_utils import extract_payroll_sheet, payroll_sheet_candidates

        workbook = self.build_minimal_payroll_named_workbook()
        file_path = os.path.join(self.temp_dir.name, "minimal_aliases.xlsx")
        with open(file_path, "wb") as file:
            file.write(workbook.read())

        candidates = payroll_sheet_candidates(file_path)
        extraction = extract_payroll_sheet(file_path, "May Staff Wages")

        self.assertIn("May Staff Wages", [candidate["sheet_name"] for candidate in candidates])
        self.assertEqual(extraction["mapping"]["s/n"], "staff_id")
        self.assertEqual(extraction["mapping"]["officer"], "full_name")
        self.assertEqual(extraction["mapping"]["monthly salary"], "basic_salary")
        self.assertEqual(extraction["mapping"]["gross earnings"], "gross_pay")
        self.assertEqual(extraction["mapping"]["income tax"], "paye")
        self.assertEqual(extraction["mapping"]["ssnit (employee)"], "ssnit")
        self.assertEqual(extraction["mapping"]["amount payable"], "net_pay")
        self.assertEqual(len(extraction["mapped_rows"]), 1)

    def test_client_sheet_matching_uses_token_overlap(self):
        from app.excel_utils import match_client_sheet

        self.assertEqual(
            match_client_sheet("ACS/GMT Shipping", ["ACS-GMT Staff Wages"]),
            "ACS-GMT Staff Wages",
        )

    def test_money_is_formatted_as_comma_separated_ghana_cedis(self):
        self.assertEqual(format_ghana_cedis(1234567.5), "GH₵ 1,234,567.50")

    def test_dashboard_requires_authentication(self):
        response = self.client.get("/dashboard")

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers["Location"])

    def test_admin_db_health_page_reports_database_counts(self):
        self.login_admin()
        response = self.client.get("/admin/db-health")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Database Type", response.data)
        self.assertIn(b"DATABASE_URL Detected", response.data)
        self.assertIn(b"Payroll Runs", response.data)
        self.assertNotIn(b"password", response.data.lower())

    def test_db_health_json_reports_connection_without_secret(self):
        self.login_admin()
        response = self.client.get("/db-health")

        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data["status"], "connected")
        self.assertIn(data["database_type"], {"SQLite", "PostgreSQL", "Other"})
        self.assertEqual(
            data["uri_prefix"],
            self.app.config["SQLALCHEMY_DATABASE_URI"].split(":", 1)[0] + "://",
        )
        self.assertNotIn("password", response.data.decode().lower())

    def test_setup_local_db_creates_database_when_missing(self):
        from setup_local_db import ensure_database_exists

        executed = []

        class FakeCursor:
            def __enter__(self):
                return self

            def __exit__(self, exc_type, exc, traceback):
                return False

            def execute(self, statement, params=None):
                executed.append((statement, params))

            def fetchone(self):
                return None

        class FakeConnection:
            def __init__(self):
                self.autocommit = False

            def cursor(self):
                return FakeCursor()

            def close(self):
                executed.append(("closed", None))

        def connect(**kwargs):
            executed.append(("connect", kwargs))
            return FakeConnection()

        ensure_database_exists(connect, db_name="chrisnat_payroll")

        self.assertIn(("connect", {"dbname": "postgres"}), executed)
        self.assertIn(
            ("CREATE DATABASE chrisnat_payroll", None),
            executed,
        )

    def test_employee_database_section_is_removed(self):
        self.login_admin()

        for path in [
            "/employee-database",
            "/employee-database/employees/new",
            "/employees",
            "/employees/export",
        ]:
            with self.subTest(path=path):
                self.assertEqual(self.client.get(path).status_code, 404)

    def test_phase2_and_phase3_routes_are_archived_from_active_mvp(self):
        self.login_admin()

        for path in [
            "/operations-dashboard",
            "/assignments",
            "/attendance",
            "/cleaning-jobs",
            "/client-portal",
            "/invoices",
            "/goods-orders",
            "/products",
            "/proposals",
        ]:
            with self.subTest(path=path):
                self.assertEqual(self.client.get(path).status_code, 404)

    def test_operations_supervisor_seed_user_exists(self):
        with self.app.app_context():
            user = User.query.filter_by(email="operations@chrisnat.local").first()

        self.assertIsNotNone(user)
        self.assertEqual(user.role, "operations_supervisor")
        self.assertTrue(user.check_password("password123"))

    def test_worker_stats_use_unique_worker_identity(self):
        rows = [
            {"staff_id": "CN-001", "full_name": "Kwame Mensah"},
            {"staff_id": "CN-001", "full_name": "Kwame Mensah"},
            {"staff_id": "", "full_name": "Ama Serwaa"},
            {"staff_id": "", "full_name": "Ama Serwaa"},
            {"staff_id": "", "full_name": ""},
        ]

        stats = calculate_worker_stats(rows)

        self.assertEqual(stats["total_rows"], 4)
        self.assertEqual(stats["total_unique_workers"], 2)
        self.assertEqual(stats["duplicate_count"], 2)

    def test_phase2_upload_creates_import_batch_preview_record(self):
        self.login_admin()
        from app.models import ImportBatch

        with self.app.app_context():
            client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            client_id = client_company.id

        response = self.client.post(
            "/payroll/runs",
            data={
                "client_company_id": str(client_id),
                "month": "February",
                "year": "2101",
                "payroll_file": (self.build_offset_phase2_workbook(), "phase2.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            batch = ImportBatch.query.filter_by(original_filename="phase2.xlsx").first()
            self.assertIsNotNone(batch)
            self.assertEqual(batch.status, "Previewed")
            self.assertEqual(batch.total_rows, 2)
            self.assertEqual(batch.total_workers, 1)
            self.assertEqual(batch.gross_total, 2400.0)
            self.assertTrue(batch.payload_json)

        import_id = response.headers["Location"].rstrip("/").split("/")[-1]
        session_path = os.path.join(self.app.config["IMPORT_SESSION_FOLDER"], f"{import_id}.json")
        self.assertFalse(os.path.exists(session_path))

    def test_upload_uses_matching_client_sheet_in_multisheet_workbook(self):
        self.login_admin()
        from app.models import ImportBatch

        with self.app.app_context():
            client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            client_id = client_company.id

        response = self.client.post(
            "/payroll/runs",
            data={
                "client_company_id": str(client_id),
                "month": "May",
                "year": "2101",
                "payroll_file": (self.build_multisheet_stress_workbook(), "stress_test.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            batch = ImportBatch.query.filter_by(original_filename="stress_test.xlsx").first()
            self.assertIsNotNone(batch)
            self.assertEqual(batch.total_rows, 3)
            self.assertEqual(batch.total_workers, 2)
            self.assertEqual(batch.gross_total, 4050.0)

        preview_response = self.client.get(response.headers["Location"])
        self.assertEqual(preview_response.status_code, 200)
        self.assertIn(b"MSC_Ghana_Ltd", preview_response.data)
        self.assertIn(b"Detected Header Row", preview_response.data)

    def test_upload_rejects_workbook_with_no_valid_payroll_rows(self):
        self.login_admin()
        from app.models import ImportBatch

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Stress_Test_Guide"
        sheet.append(["Guide only"])
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        with self.app.app_context():
            client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            client_id = client_company.id
            original_batches = ImportBatch.query.count()

        response = self.client.post(
            "/payroll/runs",
            data={
                "client_company_id": str(client_id),
                "month": "June",
                "year": "2101",
                "payroll_file": (stream, "empty_stress.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"No valid payroll rows were found", response.data)
        with self.app.app_context():
            self.assertEqual(ImportBatch.query.count(), original_batches)

    def test_multi_client_upload_previews_and_confirms_matched_client_runs(self):
        self.login_admin()
        from app.models import ImportBatch

        response = self.client.post(
            "/payroll/runs",
            data={
                "import_mode": "multi_client",
                "month": "April",
                "year": "2101",
                "payroll_file": (self.build_multiclient_workbook(), "multi_client.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        preview_response = self.client.get(response.headers["Location"])
        self.assertEqual(preview_response.status_code, 200)
        self.assertIn(b"Confirm and Create Payroll Runs", preview_response.data)
        self.assertIn(b"MSC_Ghana_Ltd", preview_response.data)
        self.assertIn(b"Stellar_Logistics", preview_response.data)
        self.assertIn(b"Unknown_Client_Payroll", preview_response.data)

        import_id = response.headers["Location"].rstrip("/").split("/")[-1]
        confirm_response = self.client.post(f"/payroll/confirm/{import_id}", follow_redirects=True)
        self.assertEqual(confirm_response.status_code, 200)

        with self.app.app_context():
            batch = db.session.get(ImportBatch, int(import_id))
            self.assertEqual(batch.status, "Imported")
            msc = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            stellar = ClientCompany.query.filter_by(name="Stellar Logistics").first()
            self.assertIsNotNone(PayrollRun.query.filter_by(client_company_id=msc.id, month="April", year=2101).first())
            self.assertIsNotNone(PayrollRun.query.filter_by(client_company_id=stellar.id, month="April", year=2101).first())

    def test_multi_client_upload_splits_consolidated_sheet_by_client_column(self):
        self.login_admin()

        response = self.client.post(
            "/payroll/runs",
            data={
                "import_mode": "multi_client",
                "month": "July",
                "year": "2101",
                "payroll_file": (self.build_consolidated_workbook(), "consolidated.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 302)
        preview_response = self.client.get(response.headers["Location"])
        self.assertEqual(preview_response.status_code, 200)
        self.assertIn(b"Consolidated_Mixed_Clients", preview_response.data)
        self.assertIn(b"MSC Ghana Ltd", preview_response.data)
        self.assertIn(b"Stellar Logistics", preview_response.data)

        import_id = response.headers["Location"].rstrip("/").split("/")[-1]
        confirm_response = self.client.post(f"/payroll/confirm/{import_id}", follow_redirects=True)
        self.assertEqual(confirm_response.status_code, 200)

        with self.app.app_context():
            msc = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            stellar = ClientCompany.query.filter_by(name="Stellar Logistics").first()
            msc_run = PayrollRun.query.filter_by(client_company_id=msc.id, month="July", year=2101).first()
            stellar_run = PayrollRun.query.filter_by(client_company_id=stellar.id, month="July", year=2101).first()
            self.assertIsNotNone(msc_run)
            self.assertIsNotNone(stellar_run)
            self.assertEqual(PayrollItem.query.filter_by(payroll_run_id=msc_run.id).count(), 1)
            self.assertEqual(PayrollItem.query.filter_by(payroll_run_id=stellar_run.id).count(), 1)

    def test_duplicate_payroll_requires_replacement_confirmation(self):
        self.login_admin()
        with self.app.app_context():
            client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            client_id = client_company.id
            # Seed a colliding run so the duplicate guard is exercised regardless of
            # the real-world date. The demo seed only creates a run for the *current*
            # month, so without this the test only passed when run in May 2026.
            db.session.add(
                PayrollRun(client_company_id=client_id, month="May", year=2026, status="Draft")
            )
            db.session.commit()
            original_count = PayrollRun.query.filter_by(
                client_company_id=client_id,
                month="May",
                year=2026,
            ).count()

        upload_response = self.client.post(
            "/payroll/runs",
            data={
                "client_company_id": str(client_id),
                "month": "May",
                "year": "2026",
                "payroll_file": (self.build_import_workbook(), "duplicate.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        import_id = upload_response.headers["Location"].rstrip("/").split("/")[-1]
        response = self.client.post(f"/payroll/confirm/{import_id}", follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"confirm replacement", response.data.lower())
        with self.app.app_context():
            new_count = PayrollRun.query.filter_by(
                client_company_id=client_id,
                month="May",
                year=2026,
            ).count()
            self.assertEqual(new_count, original_count)

    def test_main_pages_render_for_admin(self):
        response = self.login_admin()
        self.assertEqual(response.status_code, 200)

        for path in [
            "/dashboard",
            "/clients",
            "/payroll/runs",
            "/payslip",
            "/audit",
        ]:
            with self.subTest(path=path):
                self.assertEqual(self.client.get(path).status_code, 200)

    def test_removed_finance_report_and_upload_routes_are_inactive(self):
        self.login_admin()

        for path in [
            "/payroll/upload",
            "/accounts/",
            "/accounts/vouchers",
            "/accounts/remittances",
            "/accounts/expenses",
            "/reports",
        ]:
            with self.subTest(path=path):
                self.assertEqual(self.client.get(path).status_code, 404)

    def test_navigation_is_clean_payroll_mvp(self):
        response = self.login_admin()

        self.assertEqual(response.status_code, 200)
        for text in [b"Dashboard", b"Client Companies", b"Payroll Runs", b"Payslip", b"Audit"]:
            self.assertIn(text, response.data)
        for text in [
            b"Employee Database",
            b"Upload Payroll Excel",
            b"Accounts Dashboard",
            b"Payment Vouchers",
            b"Remittances",
            b"Reports",
        ]:
            self.assertNotIn(text, response.data)

    def test_dashboard_has_month_filter_sparkbars_and_action_queue(self):
        self.login_admin()
        response = self.client.get("/dashboard?month=May&year=2026")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"dashboard-controls", response.data)
        # Cost Signal column renders a sparkbar when there is cost data, or a
        # "no data" placeholder otherwise — assert the column itself is present.
        self.assertIn(b"Cost Signal", response.data)
        self.assertIn(b"Approval Queue", response.data)
        self.assertIn(b"No run submitted", response.data)

    def test_health_endpoint_is_available_for_render(self):
        response = self.client.get("/health")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["status"], "ok")

    def test_payroll_approval_queue_filter_is_server_side(self):
        self.login_admin()
        response = self.client.get("/payroll/runs?status=needs_approval")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Approval queue", response.data)
        self.assertNotIn(b"Approved</span>", response.data)

    def test_approval_updates_status_and_logs_audit_without_finance_records(self):
        self.login_admin()
        with self.app.app_context():
            payroll_run = PayrollRun.query.first()
            run_id = payroll_run.id

        response = self.client.post(f"/payroll/runs/{run_id}/approve", follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        with self.app.app_context():
            payroll_run = db.session.get(PayrollRun, run_id)
            self.assertEqual(payroll_run.status, "Approved")
            self.assertIsNone(payroll_run.voucher)
            self.assertEqual(len(payroll_run.remittances), 0)
            self.assertIsNotNone(
                AuditTrail.query.filter_by(
                    action="Payroll approval",
                    related_record_type="PayrollRun",
                    related_record_id=run_id,
                ).first()
            )

    def test_md_has_admin_clearance_for_admin_pages(self):
        response = self.login_md()
        self.assertEqual(response.status_code, 200)

        self.assertEqual(self.client.get("/clients/add").status_code, 200)
        self.assertEqual(self.client.get("/payroll/runs").status_code, 200)
        self.assertEqual(self.client.get("/payslip").status_code, 200)
        self.assertEqual(self.client.get("/audit").status_code, 200)

    def test_payroll_workflow_uses_pending_approval_and_audit(self):
        self.login_admin()

        with self.app.app_context():
            payroll_run = PayrollRun(
                month="March",
                year=2101,
                status="Draft",
                created_by=User.query.filter_by(email="admin@chrisnat.local").first().id,
                client_company_id=ClientCompany.query.filter_by(name="MSC Ghana Ltd").first().id,
                total_workers=1,
                total_gross_pay=1200,
                total_deductions=100,
                total_net_pay=1100,
                total_paye=50,
                total_ssnit=30,
            )
            db.session.add(payroll_run)
            db.session.commit()
            run_id = payroll_run.id

        self.assertEqual(
            self.client.post(f"/payroll/runs/{run_id}/submit-for-approval", follow_redirects=True).status_code,
            200,
        )
        with self.app.app_context():
            run = db.session.get(PayrollRun, run_id)
            self.assertEqual(run.status, "Pending Approval")
            self.assertIsNone(run.reviewed_by)
            self.assertIsNone(run.reviewed_at)
        self.assertEqual(
            self.client.post(f"/payroll/runs/{run_id}/approve", follow_redirects=True).status_code,
            200,
        )

        with self.app.app_context():
            run = db.session.get(PayrollRun, run_id)
            self.assertEqual(run.status, "Approved")
            self.assertIsNone(run.voucher)
            self.assertGreaterEqual(AuditTrail.query.filter_by(related_record_type="PayrollRun", related_record_id=run_id).count(), 2)

    def test_accounts_can_mark_approved_payroll_processed(self):
        self.login_admin()
        with self.app.app_context():
            payroll_run = PayrollRun.query.first()
            run_id = payroll_run.id
        self.client.post(f"/payroll/runs/{run_id}/approve", follow_redirects=True)
        self.client.get("/logout")
        self.client.post(
            "/login",
            data={"email": "accounts@chrisnat.local", "password": "password123"},
            follow_redirects=True,
        )

        response = self.client.post(f"/payroll/runs/{run_id}/mark-paid", follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        with self.app.app_context():
            run = db.session.get(PayrollRun, run_id)
            self.assertEqual(run.status, "Processed")

    def test_audit_page_records_expenses(self):
        self.login_md()
        with self.app.app_context():
            client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            payroll_run = PayrollRun.query.filter_by(client_company_id=client_company.id).first()
            client_id = client_company.id
            run_id = payroll_run.id

        response = self.client.post(
            "/audit/expenses",
            data={
                "title": "Client visit transport",
                "expense_date": "2026-05-19",
                "category": "Transport",
                "description": "MD approved client visit transport",
                "amount": "125.50",
                "client_company_id": str(client_id),
                "payroll_run_id": str(run_id),
                "recorded_by": "Managing Director",
            },
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Expense recorded.", response.data)
        self.assertIn(b"Client visit transport", response.data)
        self.assertIn(b"Tuesday", response.data)
        with self.app.app_context():
            expense = Expense.query.filter_by(title="Client visit transport").first()
            self.assertIsNotNone(expense)
            self.assertEqual(expense.payroll_run_id, run_id)
            self.assertIsNotNone(AuditTrail.query.filter_by(action="Expense recorded").first())

    def test_payslip_pdf_service_creates_pdf_file(self):
        with self.app.app_context():
            item = PayrollItem.query.first()
            file_path = generate_payslip_pdf(item, self.app.config["EXPORT_FOLDER"])

        self.assertTrue(os.path.exists(file_path))
        with open(file_path, "rb") as file:
            self.assertEqual(file.read(4), b"%PDF")

    def test_authenticated_user_can_download_individual_payslip(self):
        self.login_admin()
        with self.app.app_context():
            item = PayrollItem.query.first()
            item_id = item.id

        response = self.client.get(f"/payroll/items/{item_id}/payslip")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["Content-Type"], "application/pdf")
        self.assertTrue(response.data.startswith(b"%PDF"))
        response.close()

    def test_payroll_detail_uses_excel_grid_and_cedis_format(self):
        self.login_admin()
        with self.app.app_context():
            payroll_run = PayrollRun.query.first()
            run_id = payroll_run.id

        response = self.client.get(f"/payroll/runs/{run_id}")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"excel-grid", response.data)
        self.assertIn("GH₵ 8,044.50".encode("utf-8"), response.data)

    def test_inactive_client_status_is_red(self):
        self.login_admin()
        with self.app.app_context():
            client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            client_company.status = "Inactive"
            db.session.commit()

        response = self.client.get("/clients")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"text-bg-danger", response.data)

    def test_confirmed_import_creates_employee_records(self):
        self.login_admin()
        with self.app.app_context():
            client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            client_id = client_company.id

        upload_response = self.client.post(
            "/payroll/runs",
            data={
                "client_company_id": str(client_id),
                "month": "January",
                "year": "2101",
                "payroll_file": (self.build_import_workbook(), "new_workers.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        import_id = upload_response.headers["Location"].rstrip("/").split("/")[-1]
        response = self.client.post(f"/payroll/confirm/{import_id}", follow_redirects=True)

        self.assertEqual(response.status_code, 200)
        with self.app.app_context():
            employee = Employee.query.filter_by(staff_id="NEW-101", client_company_id=client_id).first()
            self.assertIsNotNone(employee)
            self.assertEqual(employee.full_name, "New Import Worker")
            self.assertEqual(employee.ssnit_number, "SSNIT-NEW-101")

    def test_payroll_runs_filter_by_client_tab(self):
        self.login_admin()
        with self.app.app_context():
            msc = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            stellar = ClientCompany.query.filter_by(name="Stellar Logistics").first()
            stellar_run = PayrollRun(
                month="January",
                year=2101,
                status="Draft",
                created_by=User.query.filter_by(email="admin@chrisnat.local").first().id,
                client_company_id=stellar.id,
                total_workers=1,
                source_filename="stellar.xlsx",
                total_net_pay=500,
            )
            db.session.add(stellar_run)
            db.session.commit()
            msc_id = msc.id

        response = self.client.get(f"/payroll/runs?client_id={msc_id}")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"MSC Ghana Ltd", response.data)
        self.assertNotIn(b"stellar.xlsx", response.data)

    def test_payroll_runs_page_contains_embedded_upload_form(self):
        self.login_admin()
        response = self.client.get("/payroll/runs")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Create Payroll Run", response.data)
        self.assertIn(b'name="payroll_file"', response.data)
        self.assertIn(b'name="client_company_id"', response.data)
        self.assertIn(b"excel-grid", response.data)
        self.assertNotIn(b'href="/payroll/upload"', response.data)

    def test_payslip_module_selects_client_run_and_generates_view(self):
        self.login_admin()
        with self.app.app_context():
            payroll_run = PayrollRun.query.first()
            run_id = payroll_run.id
            client_id = payroll_run.client_company_id
            item_ids = [item.id for item in payroll_run.items[:2]]

        index_response = self.client.get(f"/payslip?client_id={client_id}&run_id={run_id}")
        self.assertEqual(index_response.status_code, 200)
        self.assertIn(b"Payslip", index_response.data)
        self.assertIn(b"excel-grid", index_response.data)

        generate_response = self.client.post(
            "/payslip/generate",
            data={"payroll_item_ids": [str(item_id) for item_id in item_ids]},
            follow_redirects=True,
        )

        self.assertEqual(generate_response.status_code, 200)
        self.assertIn(b"Generated Payslips", generate_response.data)
        self.assertIn(b"Payroll run reference", generate_response.data)
        with self.app.app_context():
            self.assertGreaterEqual(AuditTrail.query.filter_by(action="Payslip generated").count(), len(item_ids))

    @unittest.skip("Proposal drafting is archived while the active MVP is reverted to payroll Phase 1.")
    def test_admin_can_create_proposal_draft_for_client(self):
        self.login_admin()
        with self.app.app_context():
            client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            client_id = client_company.id

        response = self.client.post(
            "/proposals",
            data={
                "client_company_id": str(client_id),
                "title": "Payroll Outsourcing Proposal",
                "service_summary": "Monthly payroll processing and statutory reporting.",
                "proposed_amount": "2500",
                "status": "Draft",
            },
            follow_redirects=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Payroll Outsourcing Proposal", response.data)
        from app.models import Proposal

        with self.app.app_context():
            self.assertIsNotNone(Proposal.query.filter_by(title="Payroll Outsourcing Proposal").first())

    def test_client_detail_shows_phase2_payroll_dashboard_metrics(self):
        self.login_admin()
        with self.app.app_context():
            client_company = ClientCompany.query.filter_by(name="MSC Ghana Ltd").first()
            client_id = client_company.id

        response = self.client.get(f"/clients/{client_id}")

        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Current Month Payroll Cost", response.data)
        self.assertIn(b"Previous Month Payroll Cost", response.data)
        self.assertIn(b"Validation Warnings", response.data)

    def test_audit_page_uses_excel_grid_and_audit_name(self):
        self.login_md()
        response = self.client.get("/audit")

        self.assertEqual(response.status_code, 200)
        # Page is now "Expenses & Audit" and leads with the read-only audit trail.
        self.assertIn(b"Expenses &amp; Audit", response.data)
        self.assertIn(b"Audit Trail", response.data)
        self.assertIn(b"excel-grid", response.data)
        self.assertIn(b"Day", response.data)


if __name__ == "__main__":
    unittest.main()
