"""GATE 6 — Monthly template generation & end-to-end hardening.

The generated per-company template round-trips back through Phase 3, and the
full monthly cycle (seed → template → thin upload → compute → validate → export)
runs end-to-end for DZ, with audit-trail entries and workbook preservation.
"""
import os
import tempfile
import unittest
from datetime import date

os.environ["DATABASE_URL"] = "sqlite:///:memory:"

import openpyxl

from app import create_app, db
from app.models import (
    AuditTrail,
    ClientCompany,
    Employee,
    PayrollRun,
    StatutoryRate,
)
from app.raw_engine.cleaning import normalise_element, normalise_emp_id
from app.raw_engine.exports.service import generate_run_exports
from app.raw_engine.mapping import ELEMENT_SET
from app.raw_engine.run import compute_seed_month
from app.raw_engine.seed import parse_rich_workbook
from app.raw_engine.store import persist_seed, write_payroll_items
from app.raw_engine.template import generate_monthly_template, seeded_element_codes
from app.raw_engine.thin import join_and_compute, parse_thin_workbook
from app.raw_engine.validation import validate_run

DZ_FIXTURE = os.path.join(
    os.path.dirname(__file__), "fixtures", "DZ-PAYROLL JAN 2026.xlsx"
)
_LABEL_TO_CODE = {normalise_element(l): c for c, l, _cat in ELEMENT_SET}


@unittest.skipUnless(
    os.path.exists(DZ_FIXTURE),
    "Decrypted DZ specimen missing — see tests/test_raw_engine_seed.py.",
)
class MonthlyTemplateTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app()
        self.app.config["TESTING"] = True
        self.ctx = self.app.app_context()
        self.ctx.push()
        self.client = ClientCompany(name="DZVANS COMPANY LIMITED", status="Active")
        db.session.add(self.client)
        db.session.commit()
        self.context = parse_rich_workbook(DZ_FIXTURE, self.client.id)
        persist_seed(self.context, preserve=False)
        self.rate = StatutoryRate.active_for(date(2026, 1, 1))
        self.tmp = tempfile.mkdtemp()

    def tearDown(self):
        db.session.rollback()
        self.ctx.pop()

    def _run(self):
        run = PayrollRun(
            month="January", year=2026, status="Draft",
            upload_type="raw", client_company_id=self.client.id,
        )
        db.session.add(run)
        db.session.commit()
        return run

    def _emp(self, name):
        return Employee.query.filter_by(
            client_company_id=self.client.id, full_name=name
        ).one()

    # ---- template shape --------------------------------------------------

    def test_template_has_seeded_elements_and_scoped_icu_column(self):
        path = generate_monthly_template(self.client.id, self.tmp, "February", 2026)
        header = [c.value for c in openpyxl.load_workbook(path).active[1]]
        self.assertEqual(header[0], "Staff ID")
        self.assertIn("ICU Member", header)
        for _code, label, _cat in ELEMENT_SET:  # all 9 seeded element columns
            self.assertIn(label, header)

        ws = openpyxl.load_workbook(path).active
        icu_col = header.index("ICU Member") + 1
        staff_col = 1
        flags = {}
        for r in range(2, ws.max_row + 1):
            flags[normalise_emp_id(ws.cell(r, staff_col).value)] = ws.cell(r, icu_col).value
        member = self._emp("RICHARD WOODE")       # ICU member
        nonmember = self._emp("GEORGE AKOTO")     # non-member
        self.assertEqual(flags[normalise_emp_id(member.staff_id)], "Member")
        self.assertIn(flags[normalise_emp_id(nonmember.staff_id)], (None, ""))

    # ---- round-trip (blank template) ------------------------------------

    def test_generated_template_round_trips_through_phase3(self):
        path = generate_monthly_template(self.client.id, self.tmp, "February", 2026)
        inputs, _warn = parse_thin_workbook(path)
        self.assertEqual(len(inputs), 181)

        result = join_and_compute(inputs, self._run(), self.rate)
        self.assertEqual(result.blocked, [])
        self.assertEqual(len(result.payslips), 181)

        # Blank template => zero hours: an hourly worker costs 0 (NOT the stale
        # seed basic), a salaried admin still gets the flat basic.
        woode = result.payslips[normalise_emp_id(self._emp("RICHARD WOODE").staff_id)]
        self.assertEqual(woode.basic_wage, 0.0)
        self.assertEqual(woode.gross_pay, 0.0)
        george = result.payslips[normalise_emp_id(self._emp("GEORGE AKOTO").staff_id)]
        self.assertAlmostEqual(george.basic_wage, 1800.0, delta=0.001)

    # ---- full monthly cycle ---------------------------------------------

    def test_full_monthly_cycle_end_to_end(self):
        """seed → template → fill → thin upload → compute → validate → export.
        Filling the generated template with the month's hours reproduces the
        Phase 2 payroll."""
        phase2 = compute_seed_month(self.context, self.rate)

        template_path = generate_monthly_template(self.client.id, self.tmp, "January", 2026)
        self._fill_template_from_seed(template_path)

        inputs, _ = parse_thin_workbook(template_path)
        run = self._run()
        result = join_and_compute(inputs, run, self.rate)
        self.assertEqual(result.blocked, [])

        # Reproduces Phase 2 per employee.
        for staff_id, slip2 in phase2.items():
            slip = result.payslips[staff_id]
            self.assertAlmostEqual(slip.gross_pay, slip2.gross_pay, delta=0.001)
            self.assertAlmostEqual(slip.net_pay, slip2.net_pay, delta=0.001)

        # Persist, validate, export.
        write_payroll_items(run, result.payslips)
        membership = {
            e.staff_id: bool(e.icu_member)
            for e in Employee.query.filter_by(client_company_id=self.client.id)
        }
        report = validate_run(run.items, membership)
        self.assertTrue(report.is_savable, msg=[b.message for b in report.blocks])

        exports = generate_run_exports(run, self.tmp)
        self.assertTrue(exports["routing"]["complete"])
        for path in exports["files"].values():
            self.assertTrue(os.path.exists(path))

    # ---- hardening: audit + preservation --------------------------------

    def test_audit_entries_on_seed_and_compute(self):
        run = self._run()
        write_payroll_items(run, compute_seed_month(self.context, self.rate))
        actions = {a.action for a in AuditTrail.query.all()}
        self.assertIn("Raw payroll seed confirmed", actions)   # from setUp's persist_seed
        self.assertIn("Raw payroll computed", actions)

    def test_workbook_bytes_preserved_on_seed(self):
        fresh = ClientCompany(name="DZVANS PRESERVE TEST", status="Active")
        db.session.add(fresh)
        db.session.commit()
        context = parse_rich_workbook(DZ_FIXTURE, fresh.id)
        result = persist_seed(context, source_path=DZ_FIXTURE, preserve=True)
        self.assertTrue(result.workbook_preserved_to)
        self.assertTrue(os.path.exists(result.workbook_preserved_to))

    def test_seeded_element_codes_derived_from_data(self):
        codes = seeded_element_codes(self.client.id)
        self.assertIn("NORMAL", codes)
        self.assertIn("SAT_OT", codes)
        # ordered as in ELEMENT_SET
        self.assertLess(codes.index("NORMAL"), codes.index("SAT_OT"))

    # ---- helper ----------------------------------------------------------

    def _fill_template_from_seed(self, path):
        """Fill the generated (zero) template with each worker's seed-month
        hours and adjustments — the client's monthly edit."""
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        col_code = {}
        adj_col = {}
        for idx, h in enumerate(header, start=1):
            hn = normalise_element(h)
            if hn in _LABEL_TO_CODE:
                col_code[idx] = _LABEL_TO_CODE[hn]
            elif hn in ("PROD/BONUS ALLOWANCE",):
                adj_col[idx] = "bonus"
            elif hn == "LOAN":
                adj_col[idx] = "loan"
            elif hn == "WELFARE":
                adj_col[idx] = "welfare"
            elif hn == "OTHER DEDUCTION":
                adj_col[idx] = "other_deduction"
            elif hn == "PAY DIFFERENCE":
                adj_col[idx] = "pay_difference"

        by_staff = {e.staff_id: e for e in self.context.employees}
        for r in range(2, ws.max_row + 1):
            emp = by_staff.get(normalise_emp_id(ws.cell(r, 1).value))
            if not emp:
                continue
            for idx, code in col_code.items():
                ws.cell(r, idx, emp.raw_hours.get(code, 0))
            for idx, field in adj_col.items():
                ws.cell(r, idx, getattr(emp, field))
        wb.save(path)


if __name__ == "__main__":
    unittest.main()
