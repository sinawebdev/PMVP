"""GATE 4 — Raw Hours Engine validation layer.

Each rule fires on a crafted bad input and a clean input passes: hours
reconciliation (anchored to the thin consolidation specimen's 15,945.66 grand
total), ICU tie-out, non-member-with-ICU flag, unknown-Staff-ID block, and
recompute drift.
"""
import os
import unittest
from datetime import date

os.environ["DATABASE_URL"] = "sqlite:///:memory:"

import openpyxl

from app import create_app, db
from app.models import ClientCompany, Employee, PayrollRun, StatutoryRate
from app.raw_engine.consolidation import consolidate_hours
from app.raw_engine.icu_distribution import distribute_union_dues
from app.raw_engine.matrix import build_matrix
from app.raw_engine.run import compute_seed_month
from app.raw_engine.seed import parse_rich_workbook
from app.raw_engine.store import persist_seed, write_payroll_items
from app.raw_engine.validation import (
    BLOCK,
    FLAG,
    WARNING,
    ValidationReport,
    block_unknown_staff,
    check_icu_tie_out,
    flag_nonmember_icu,
    reconcile_hours,
    recompute_drift,
    validate_run,
)

DZ_FIXTURE = os.path.join(
    os.path.dirname(__file__), "fixtures", "DZ-PAYROLL JAN 2026.xlsx"
)
THIN_FIXTURE = os.path.join(
    os.path.dirname(__file__), "fixtures", "Rawdata to work on data_consolidation.xlsx"
)
ANCHOR_TOTAL = 15945.66


class _Item:
    """Minimal PayrollItem stand-in for the pure orchestrator tests."""

    def __init__(self, staff_id, icu_dues=0.0, gross_pay=0.0, net_pay=0.0):
        self.staff_id = staff_id
        self.icu_dues = icu_dues
        self.gross_pay = gross_pay
        self.net_pay = net_pay


class ValidationRuleTests(unittest.TestCase):
    # ---- hours reconciliation -------------------------------------------

    def test_hours_reconciliation_passes_and_is_case_insensitive(self):
        rows = [("DCL1", "Normal hours", 100), ("DCL1", "NORMAL HOURS", 50)]
        consolidated = consolidate_hours(rows)  # merges case-insensitively -> 150
        self.assertIsNone(reconcile_hours(rows, consolidated))
        self.assertEqual(build_matrix(consolidated).grand_total, 150)

    def test_hours_reconciliation_blocks_on_dropped_row(self):
        rows = [("DCL1", "NORMAL", 100), ("", "NORMAL", 50)]  # blank staff dropped
        consolidated = consolidate_hours(rows)
        issue = reconcile_hours(rows, consolidated)
        self.assertIsNotNone(issue)
        self.assertEqual(issue.severity, BLOCK)
        self.assertEqual(issue.detail["raw_total"], 150.0)
        self.assertEqual(issue.detail["consolidated_total"], 100.0)

    # ---- ICU tie-out -----------------------------------------------------

    def test_icu_tie_out_passes_when_cascade_ties(self):
        self.assertIsNone(check_icu_tie_out(1000.00))
        # Awkward total still ties (remainder correction).
        self.assertIsNone(check_icu_tie_out(100.01))

    def test_icu_tie_out_blocks_on_remittance_mismatch(self):
        # Export remits a distribution built from a stale total (a member missing).
        stale = distribute_union_dues(900.00)
        issue = check_icu_tie_out(1000.00, distribution=stale)
        self.assertIsNotNone(issue)
        self.assertEqual(issue.severity, BLOCK)

    def test_union_cascade_shares_and_totals(self):
        d = distribute_union_dues(1000.00)
        self.assertAlmostEqual(d.union, 500.00, delta=0.001)
        self.assertAlmostEqual(d.edfund, 500.00, delta=0.001)
        self.assertAlmostEqual(d.icu_accra, 375.00, delta=0.001)  # 75% of union
        self.assertAlmostEqual(d.local, 100.00, delta=0.001)      # 20% of union
        self.assertAlmostEqual(d.icu_tema, 25.00, delta=0.001)    # 5% of union
        self.assertAlmostEqual(d.icu_edac, 400.00, delta=0.001)   # 80% of edfund
        self.assertAlmostEqual(d.dcl_eef, 100.00, delta=0.001)    # 20% of edfund
        self.assertAlmostEqual(d.total_payout, 1000.00, delta=0.001)

    # ---- non-member ICU flag --------------------------------------------

    def test_nonmember_with_icu_is_flagged_member_is_not(self):
        issues = flag_nonmember_icu([
            ("M1", 18.94, True),    # member with dues — fine
            ("N1", 12.00, False),   # non-member with dues — flag
            ("N2", 0.0, False),     # non-member, no dues — fine
        ])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].severity, FLAG)
        self.assertEqual(issues[0].detail["staff_id"], "N1")

    # ---- unknown staff block --------------------------------------------

    def test_unknown_staff_id_blocks(self):
        issues = block_unknown_staff([
            {"staff_id": "GHOST", "reason": "Unknown Staff ID — seed first."}
        ])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].severity, BLOCK)

    # ---- recompute drift -------------------------------------------------

    def test_recompute_drift_warns_beyond_tolerance(self):
        issues = recompute_drift([
            ("A", "net", 100.00, 100.01),   # within tol -> no warning
            ("B", "gross", 500.00, 505.00),  # drift -> warning
        ])
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].severity, WARNING)
        self.assertEqual(issues[0].detail["staff_id"], "B")

    # ---- orchestrator savable gate --------------------------------------

    def test_clean_run_is_savable_bad_run_is_not(self):
        clean = validate_run(
            [_Item("M1", icu_dues=18.94), _Item("M2", icu_dues=10.00)],
            membership={"M1": True, "M2": True},
        )
        self.assertTrue(clean.is_savable)
        self.assertEqual(clean.blocks, [])

        bad = validate_run(
            [_Item("N1", icu_dues=12.00)],
            membership={"N1": False},              # non-member with dues -> flag
            blocked=[{"staff_id": "GHOST", "reason": "unknown"}],  # -> block
        )
        self.assertFalse(bad.is_savable)
        self.assertTrue(bad.blocks)
        self.assertTrue(bad.flags)


@unittest.skipUnless(
    os.path.exists(THIN_FIXTURE),
    "Thin consolidation specimen missing from tests/fixtures/.",
)
class HoursAnchorTests(unittest.TestCase):
    def test_consolidation_reconciles_to_15945_66(self):
        ws = openpyxl.load_workbook(THIN_FIXTURE, data_only=True)["Sheet1"]
        rows = []
        for r in range(2, ws.max_row + 1):
            staff, hours, period = (
                ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
            )
            if staff is None and hours is None:
                continue
            rows.append((staff, period, hours))  # period column is the element label

        consolidated = consolidate_hours(rows)
        matrix = build_matrix(consolidated)
        self.assertAlmostEqual(matrix.grand_total, ANCHOR_TOTAL, delta=0.01)
        # Clean file: hours reconciliation passes.
        self.assertIsNone(reconcile_hours(rows, consolidated))


@unittest.skipUnless(
    os.path.exists(DZ_FIXTURE),
    "Decrypted DZ specimen missing — see tests/test_raw_engine_seed.py.",
)
class FullRunValidationTests(unittest.TestCase):
    def setUp(self):
        self.app = create_app()
        self.app.config["TESTING"] = True
        self.ctx = self.app.app_context()
        self.ctx.push()
        self.client = ClientCompany(name="DZVANS COMPANY LIMITED", status="Active")
        db.session.add(self.client)
        db.session.commit()
        self.context = parse_rich_workbook(DZ_FIXTURE, self.client.id)
        persist_seed(self.context, preserve=False)
        self.rate = StatutoryRate.active_for(date(2026, 1, 1))

    def tearDown(self):
        db.session.rollback()
        self.ctx.pop()

    def test_full_dz_run_validates_clean_and_icu_ties(self):
        run = PayrollRun(
            month="January", year=2026, status="Draft",
            upload_type="raw", client_company_id=self.client.id,
        )
        db.session.add(run)
        db.session.commit()
        write_payroll_items(run, compute_seed_month(self.context, self.rate))

        items = run.items
        membership = {
            e.staff_id: bool(e.icu_member)
            for e in Employee.query.filter_by(client_company_id=self.client.id)
        }
        report = validate_run(items, membership)

        self.assertTrue(report.is_savable, msg=[b.message for b in report.blocks])
        self.assertEqual(report.blocks, [])
        self.assertEqual(report.flags, [])  # every ICU dues sits on a real member

        # ICU dues tie out to the union distribution exactly.
        icu_total = round(sum((i.icu_dues or 0) for i in items), 2)
        self.assertGreater(icu_total, 0)
        self.assertAlmostEqual(
            distribute_union_dues(icu_total).total_payout, icu_total, delta=0.01
        )

    def test_injected_nonmember_icu_flags_but_does_not_block(self):
        run = PayrollRun(
            month="January", year=2026, status="Draft",
            upload_type="raw", client_company_id=self.client.id,
        )
        db.session.add(run)
        db.session.commit()
        write_payroll_items(run, compute_seed_month(self.context, self.rate))

        items = run.items
        membership = {
            e.staff_id: bool(e.icu_member)
            for e in Employee.query.filter_by(client_company_id=self.client.id)
        }
        # A member carrying ICU dues, mislabelled non-member in the membership map.
        member_item = next(i for i in items if (i.icu_dues or 0) > 0)
        membership[member_item.staff_id] = False

        report = validate_run(items, membership)
        self.assertTrue(any(f.detail["staff_id"] == member_item.staff_id
                            for f in report.flags))
        self.assertTrue(report.is_savable)  # a flag never blocks


if __name__ == "__main__":
    unittest.main()
